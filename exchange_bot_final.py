from dotenv import load_dotenv
load_dotenv()
import os
import logging
import asyncio
import re
import json
import httpx
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReactionTypeEmoji, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler,
    MessageReactionHandler,
    filters, ContextTypes
)

MAIN_KB = ReplyKeyboardMarkup([[KeyboardButton('/list')]], resize_keyboard=True, is_persistent=True)

MAIN_KB = ReplyKeyboardMarkup([[KeyboardButton('/list')]], resize_keyboard=True, is_persistent=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
GROQ_API_KEY   = os.environ.get("GROQ_API_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

WAITING_RATE = 0
orders = {}
current_rate = {}
bot_active = {}  # chat_id -> bool
sessions = {}  # chat_id -> список архивных сессий
debt = {}  # chat_id -> остаток долга в USD
last_bot_msg = {}  # chat_id -> message_id последнего сообщения бота
last_list_msg = {}  # chat_id -> message_id последнего /list (для обновления вместо нового)
balance = {}  # chat_id -> остаток баланса в RUB (None = не задан)
bot_msgs = {}     # chat_id -> [последние 3 message_id, отправленные ботом]
topups = {}   # chat_id -> [{"date": str, "amount": int}] — история пополнений текущей сессии
last_bulk_close = {}  # chat_id -> [id1, id2, ...] для кнопки "Отменить закрытие всех"

def get_orders(chat_id):
    if chat_id not in orders:
        orders[chat_id] = []
    return orders[chat_id]

def next_id(chat_id):
    existing = get_orders(chat_id)
    return max((o["id"] for o in existing), default=0) + 1

def get_rate(chat_id):
    return current_rate.get(chat_id)

def fmt_usd(val):
    """Форматирует сумму USD: всегда 3 знака после запятой."""
    if val is None:
        return "0.000 USD"
    s = f"{float(val):,.3f}".replace(",", " ")
    return s + " USD"

def _balance_line(chat_id):
    bal_usd = balance.get(chat_id)
    if bal_usd is None:
        return ""
    rate = get_rate(chat_id)
    if rate:
        rub = bal_usd * rate
        rub_str = f"{int(rub):,}".replace(",", " ") + " ₽"
        bal_str = f"{rub_str} / {fmt_usd(bal_usd)}"
    else:
        bal_str = fmt_usd(bal_usd)
    return f"💰 Баланс: {bal_str}\n"

def _order_to_usd(order, default_rate):
    amt = order.get("amount") or 0
    cur = order.get("currency", "RUB")
    r = order.get("close_rate") or default_rate
    if not amt:
        return 0
    if cur in ("USDT", "USD", "USDC"):
        return amt
    if cur == "RUB" and r:
        return amt / r
    return 0

def _format_topups_block(s):
    """Возвращает строки про пополнения и итоговый баланс из архивной сессии."""
    out = []
    tps = s.get("topups") or []
    if tps:
        total = sum(t.get("amount", 0) for t in tps)
        parts = []
        for t in tps:
            a = t.get("amount", 0)
            parts.append(f"+{int(a/1000)}к")
        total_str = f"{int(total):,}".replace(",", " ") + " ₽"
        out.append(f"  💰 Пополнения: {', '.join(parts)} = {total_str}")
    fb = s.get("final_balance")
    if fb is not None:
        bal_str = f"{int(fb):,}".replace(",", " ") + " ₽"
        if fb < 0:
            out.append(f"  📉 Баланс на момент закрытия: {bal_str} (не хватило)")
        elif fb > 0:
            out.append(f"  ✅ Баланс на момент закрытия: {bal_str} (остаток)")
        else:
            out.append(f"  💰 Баланс на момент закрытия: 0 ₽")
    return out


SYSTEM_PROMPT = """Ты парсер платёжных заявок. Возвращай ТОЛЬКО валидный JSON, без markdown, без объяснений.

ПРАВИЛА:
- "Xк" / "Xk" = X*1000. Например: "11к"=11000, "100к"=100000, "1.5к"=1500.
- "Xт.р" / "Xтр" / "Xтыс" = X*1000. Например: "25т.р"=25000.
- Просто число (без суффикса) — это сумма как есть, например "5000"=5000.
- Разделители тысяч: "5 600"=5600, "32 500"=32500.
- amount возвращай ВСЕГДА как число (int), а не строку.

РЕКВИЗИТЫ:
- card: ровно 16 цифр без пробелов и тире
- phone: формат "+7XXXXXXXXXX" (10 цифр после +7). Принимай любой формат: 89..., 79..., +7 (914) 321-72-04
- card и phone — это РАЗНЫЕ поля. 10-11 цифр = phone. 16 цифр = card.

БАНК: возвращай каноничное название из текста: Сбер, Т-Банк (он же Тинькофф), ВТБ, Альфа, Озон, Райффайзен, Газпромбанк, Совкомбанк, ОТП, МКБ, ПСБ, ЮMoney (юма ни/юмани), QIWI, Почта Банк, Уралсиб, Хоум Кредит, Росбанк, Открытие, Ренессанс, Яндекс Банк, Абсолют, Авангард, АК Барс, БСП, ВБРР, Экспобанк, WB (Wildberries/Вайлбериз), Цифра банк, Русский Стандарт. Если не нашёл — null.

ПРИМЕРЫ:
"СБП ВТБ 89083132718 11к" → {"bank":"ВТБ","amount":11000,"currency":"RUB","card":null,"phone":"+79083132718"}
"Сбер 5 600 +79161234567" → {"bank":"Сбер","amount":5600,"currency":"RUB","card":null,"phone":"+79161234567"}
"2202201740186792 - 25000 сбер" → {"bank":"Сбер","amount":25000,"currency":"RUB","card":"2202201740186792","phone":null}
"Юма ни 89521683870 45000" → {"bank":"ЮMoney","amount":45000,"currency":"RUB","card":null,"phone":"+79521683870"}

ФОРМАТ ОТВЕТА (строго):
{"bank": string|null, "amount": number|null, "currency": "RUB", "card": string|null, "phone": string|null}"""

async def parse_with_groq(text):
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                json={
                    "model": "llama-3.3-70b-versatile",
                    "temperature": 0,
                    "max_tokens": 300,
                    "response_format": {"type": "json_object"},
                    "messages": [
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": re.sub(r" - ", " ", normalize_amount_spaces(text))}
                    ]
                }
            )
            raw = resp.json()["choices"][0]["message"]["content"].strip()
            logger.info(f"Groq raw: {repr(raw)}")
            raw = re.sub(r"```json|```", "", raw).strip()
            # Вырезаем только JSON объект если есть лишний текст
            match = re.search(r'\{[^{}]+\}', raw, re.DOTALL)
            if not match:
                raise ValueError("No JSON found")
            parsed = json.loads(match.group())
            logger.info(f"Groq: {parsed}")
            if parsed.get("amount") or parsed.get("card") or parsed.get("phone"):
                return parsed
    except Exception as e:
        logger.error(f"Groq error: {e}")
    return None

def parse_amount(val):
    """Конвертируем сумму: '220к'->220000, '1.5к'->1500, 220->220"""
    if val is None:
        return None
    s = str(val).strip().lower().replace(" ", "")
    if s.endswith("к") or s.endswith("k"):
        try:
            result = float(s[:-1]) * 1000
            return result if result <= 5_000_000 else None
        except:
            return None
    try:
        result = float(s)
        digits_only = re.sub(r"\D", "", s)
        # Отсекаем: больше 50млн, меньше 10, или 7+ цифр (похоже на телефон/реквизиты)
        if result > 5_000_000:
            return None
        if result < 10:
            return None
        if len(digits_only) >= 7:
            return None
        return result
    except:
        return None

def format_label(order):
    parts = []
    if order.get("bank"):
        parts.append(order["bank"])
    if order.get("amount") is not None:
        amt = int(order["amount"]) if float(order["amount"]).is_integer() else order["amount"]
        parts.append(f"{amt:,} {order.get('currency', 'RUB')}".replace(",", " "))
    if order.get("rate"):
        parts.append(f"курс {order['rate']}")
    if order.get("card"):
        parts.append(f"···{order['card'][-4:]}")
    elif order.get("phone"):
        parts.append(f"···{order['phone'][-4:]}")
    return " · ".join(parts) if parts else "Заявка"

async def setrate_start(update, ctx):
    chat_id = update.effective_chat.id
    if ctx.args:
        try:
            rate = float(ctx.args[0].replace(",", "."))
            current_rate[chat_id] = rate
            applied = 0
            for o in get_orders(chat_id):
                if o.get("rate") is None:
                    o["rate"] = rate
                    applied += 1
            extra = f"\n📌 Применён к {applied} заявкам без курса" if applied else ""
            await update.message.reply_text(f"✅ Курс установлен: {rate} ₽{extra}")
            return ConversationHandler.END
        except ValueError:
            pass
    cur = current_rate.get(chat_id)
    cur_text = f"Текущий курс: {cur} ₽\n\n" if cur else ""
    await update.message.reply_text(
        f"{cur_text}📌 Введи новый курс (например: 77.5):",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← Меню", callback_data="cancel_conv")]])
    )
    return WAITING_RATE

async def setrate_save(update, ctx):
    chat_id = update.effective_chat.id
    try:
        rate = float(update.message.text.replace(",", "."))
        current_rate[chat_id] = rate
        applied = 0
        for o in get_orders(chat_id):
            if o.get("rate") is None:
                o["rate"] = rate
                applied += 1
        extra = f"\n📌 Применён к {applied} заявкам без курса" if applied else ""
        await update.message.reply_text(f"✅ Курс установлен: {rate} ₽{extra}")
    except ValueError:
        await update.message.reply_text(
            "❌ Введи число, например: 77.5",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← Меню", callback_data="cancel_conv")]])
        )
        return WAITING_RATE
    return ConversationHandler.END

# ─── Multi-order support ──────────────────────────────────────────

BANK_MAP = {
    "т-банк": "Т-Банк", "т банк": "Т-Банк",
    "тинь": "Т-Банк", "тинёк": "Т-Банк", "тинк": "Т-Банк",
    "тинькофф": "Т-Банк", "tinkoff": "Т-Банк",
    "сбер": "Сбер", "сбербанк": "Сбер", "втб": "ВТБ", "альфа": "Альфа",
    "озон": "Озон", "ozon": "Озон", "райф": "Райффайзен",
    "райффайзен": "Райффайзен", "газпром": "Газпромбанк",
    "совком": "Совкомбанк", "отп": "ОТП", "мкб": "МКБ", "псб": "ПСБ",
    "юмани": "ЮMoney", "юmoney": "ЮMoney", "yoomoney": "ЮMoney", "yumoney": "ЮMoney", "umoney": "ЮMoney", "qiwi": "QIWI", "киви": "QIWI",
    "почта": "Почта Банк", "уралсиб": "Уралсиб",
    "хоум": "Хоум Кредит", "росбанк": "Росбанк",
    "открытие": "Открытие", "ренессанс": "Ренессанс",
    "яндекс": "Яндекс Банк", "ябанк": "Яндекс Банк",
    "абсолют": "Абсолют", "авангард": "Авангард",
    "ак барс": "АК Барс", "акбарс": "АК Барс", "бсп": "БСП",
    "вбрр": "ВБРР", "экспо": "Экспобанк", "генбанк": "Генбанк",
    "пойдём": "Пойдём", "пойдем": "Пойдём", "цифра": "Цифра банк",
    "русский стандарт": "Русский Стандарт",
    "вайлбериз": "WB", "wildberries": "WB", "wb банк": "WB",
}

_AMT_RE   = re.compile(r'(\d+[\.,]?\d*)\s*(?:т\.?\s*р\.?|тр\b|тыс)', re.IGNORECASE)
_K_RE     = re.compile(r'(\d+[\.,]?\d*)\s*[кk]\b', re.IGNORECASE)
_PHONE_RE = re.compile(r'\+\s*7[\s\-\(\)]*(\d[\s\-\(\)]*){10}|(?<!\d)8[\s\-\(\)]*(\d[\s\-\(\)]*){10}|(?<!\d)7\d{10}(?!\d)|(?<!\d)9\d{9}(?!\d)')
_CARD_RE  = re.compile(r'(?<!\d)\d{4}[ \t\-]?\d{4}[ \t\-]?\d{4}[ \t\-]?\d{4}(?!\d)')

def normalize_amount_spaces(text):
    """Склеивает числа с разделителями тысяч:
       '5 600' → '5600' (пробел), '20.000' → '20000' (точка перед 3 цифрами).
       Не трогает 4-значные группы (карты), доли (1.5к), одиночные цифры."""
    text = re.sub(
        r'\b(\d{1,3}(?:\s+\d{3})+)\b',
        lambda m: re.sub(r'\s+', '', m.group(0)),
        text
    )
    # Точка как разделитель тысяч: 20.000 → 20000, 1.234.567 → 1234567
    prev = None
    while prev != text:
        prev = text
        text = re.sub(r'(?<!\d)(\d{1,3})\.(\d{3})(?!\d)', r'\1\2', text)
    return text

def _find_bank(text):
    t = text.lower()
    # Также пробуем без пробелов: "юма ни" → "юмани", "т банк" → "тбанк"
    t_compact = re.sub(r'\s+', '', t)
    for key, name in BANK_MAP.items():
        if key in t:
            return name
        key_compact = key.replace(' ', '')
        if key_compact != key and key_compact in t_compact:
            return name
        if ' ' not in key and key in t_compact and key not in t:
            return name
    return None

def _find_phone(text):
    m = _PHONE_RE.search(text)
    if m:
        digits = re.sub(r'\D', '', m.group())
        if 10 <= len(digits) <= 11:
            return '+7' + digits[-10:]
    return None

def _find_card(text):
    m = _CARD_RE.search(text)
    if m:
        digits = re.sub(r'\D', '', m.group())
        if len(digits) == 16:
            return digits
    return None

def _find_amount(text):
    text = normalize_amount_spaces(text)
    text_nc = _CARD_RE.sub(' CARD ', text)
    text_nc = _PHONE_RE.sub(' PHONE ', text_nc)
    text_nc = re.sub(r'\s+', ' ', text_nc)
    m = _AMT_RE.search(text_nc)
    if m:
        try:
            val = float(m.group(1).replace(',', '.')) * 1000
            if 100 <= val <= 5_000_000:
                return val
        except Exception:
            pass
    m = _K_RE.search(text_nc)
    if m:
        try:
            val = float(m.group(1).replace(',', '.')) * 1000
            if 100 <= val <= 5_000_000:
                return val
        except Exception:
            pass
    for m in re.finditer(r'\b(\d{3,7})\b', text_nc):
        try:
            val = float(m.group(1))
            if 100 <= val <= 5_000_000:
                return val
        except Exception:
            pass
    return None

def parse_chunk_local(chunk):
    """Regex-парсер одной заявки (без AI)"""
    phone  = _find_phone(chunk)
    card   = _find_card(chunk) if not phone else None
    amount = _find_amount(chunk)
    bank   = _find_bank(chunk)
    return {"bank": bank, "amount": amount, "currency": "RUB", "card": card, "phone": phone}

def is_multi_order(text):
    """True если сообщение содержит несколько заявок"""
    text = normalize_amount_spaces(text)
    phones  = _PHONE_RE.findall(text)
    cards   = _CARD_RE.findall(text)
    amounts = _AMT_RE.findall(text) + _K_RE.findall(text)
    if len(phones) >= 2 or len(cards) >= 2:
        return True
    if len(amounts) >= 2 and text.count('\n') >= 2:
        return True
    # Многострочное сообщение, где ≥2 строк выглядят как заявки
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if len(lines) >= 2:
        signal = 0
        for line in lines:
            ll = line.lower()
            has_bank  = any(k in ll for k in BANK_MAP)
            has_num   = bool(re.search(r'\b\d{2,7}\b', line))
            has_kamt  = bool(_K_RE.search(line) or _AMT_RE.search(line))
            has_phone = bool(_PHONE_RE.search(line))
            has_card  = bool(_CARD_RE.search(line))
            if has_phone or has_card or has_kamt or (has_bank and has_num):
                signal += 1
        if signal >= 2:
            return True
    return False

def _split_by_anchors(text):
    """Разбивает однострочный текст с несколькими заявками по картам/телефонам"""
    cards   = [(m.start(), m.end()) for m in _CARD_RE.finditer(text)]
    phones  = [(m.start(), m.end()) for m in _PHONE_RE.finditer(text)]
    anchors = sorted(cards + phones, key=lambda x: x[0])
    if len(anchors) < 2:
        return [text]
    # Формат: anchor-first (карта/тел в начале) vs anchor-last (после суммы/банка)
    before_first = text[:anchors[0][0]]
    before_lower = before_first.lower()
    is_anchor_last = (
        bool(_AMT_RE.search(before_first)) or
        bool(_K_RE.search(before_first)) or
        any(key in before_lower for key in BANK_MAP)
    )
    chunks = []
    last = 0
    for i in range(len(anchors) - 1):
        a_end   = anchors[i][1]
        b_start = anchors[i+1][0]
        if b_start <= a_end:
            continue
        between = text[a_end:b_start]
        if is_anchor_last:
            split_pos = a_end
        else:
            spaces = list(re.finditer(r'\s+', between))
            if spaces:
                split_pos = a_end + spaces[-1].start()
            else:
                split_pos = (a_end + b_start) // 2
        chunks.append(text[last:split_pos].strip())
        last = split_pos
    chunks.append(text[last:].strip())
    return [c for c in chunks if c]

def split_order_chunks(text):
    """Разбивает сообщение с несколькими заявками на отдельные чанки.
    Правило: новая заявка начинается когда текущий чанк уже ПОЛНЫЙ
    (имеет и сумму/банк, и реквизиты), либо когда новая строка дублирует
    то что уже есть в чанке."""
    text = normalize_amount_spaces(text)  # 20.000 → 20000, 5 600 → 5600
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    if len(lines) > 1:
        line_chunks = []
        current = []
        cur_has_amt = False
        cur_has_req = False
        for line in lines:
            line_lower = line.lower()
            has_tyr   = bool(_AMT_RE.search(line) or _K_RE.search(line))
            has_bank  = any(key in line_lower for key in BANK_MAP)
            has_num   = bool(re.search(r'\b\d{4,7}\b', line))
            line_has_amt   = has_tyr or has_num or (has_bank and has_num)
            line_has_phone = bool(_PHONE_RE.search(line))
            line_has_card  = bool(_CARD_RE.search(line)) and not line_has_phone
            line_has_req   = line_has_phone or line_has_card
            start_new = False
            if current:
                if cur_has_amt and cur_has_req:
                    if line_has_amt or line_has_req:
                        start_new = True
                elif cur_has_amt and line_has_amt:
                    start_new = True
                elif cur_has_req and line_has_req:
                    start_new = True
            if start_new:
                line_chunks.append('\n'.join(current))
                current = [line]
                cur_has_amt = line_has_amt
                cur_has_req = line_has_req
            else:
                current.append(line)
                cur_has_amt = cur_has_amt or line_has_amt
                cur_has_req = cur_has_req or line_has_req
        if current:
            line_chunks.append('\n'.join(current))
        line_chunks = [c for c in line_chunks if c.strip()]
    else:
        line_chunks = [text]
    final = []
    for chunk in line_chunks:
        sub = _split_by_anchors(chunk)
        final.extend(sub)
    return final

_MULTIPLIER_RE = re.compile(
    r'(?<!\d)(\d+)\s*(?:платеж[а-я]*|раз[а-я]*|шт\.?|штук[а-я]*)?\s*\bпо\b\s*(\d+(?:[\.,]?\d+)?[кk]?)',
    re.IGNORECASE
)

def expand_multiplier(text):
    """\"N платежей по X\" -> (count, текст с заменой). Иначе (1, text)."""
    m = _MULTIPLIER_RE.search(text)
    if not m:
        return (1, text)
    count = int(m.group(1))
    if count < 2 or count > 20:
        return (1, text)
    amount_str = m.group(2)
    base_text = text[:m.start()] + amount_str + text[m.end():]
    return (count, base_text)

async def handle_multi_order(msg, ctx, chunks):
    """Создаёт N заявок из одного сообщения со списком"""
    chat_id = msg.chat_id
    added   = []
    skipped = 0
    # Раскрываем "N платежей по X" — каждый чанк может породить несколько заявок
    expanded_chunks = []
    for chunk in chunks:
        count, base = expand_multiplier(chunk)
        for _ in range(count):
            expanded_chunks.append(base)
    for chunk in expanded_chunks:
        parsed = parse_chunk_local(chunk)
        if not parsed.get("amount") and not parsed.get("card") and not parsed.get("phone"):
            skipped += 1
            continue
        order_id = next_id(chat_id)
        order = {
            "id": order_id, "label": "",
            "original_text": chunk,
            "bank": parsed.get("bank"),
            "amount": parsed.get("amount"),
            "currency": "RUB",
            "card": parsed.get("card"),
            "phone": parsed.get("phone"),
            "rate": get_rate(chat_id), "close_rate": None,
            "status": "open", "closed_by": None, "closed_at": None,
            "message_id": msg.message_id,
            "added_at": datetime.now().strftime("%H:%M"),
        }
        order["label"] = format_label(order)
        get_orders(chat_id).append(order)
        added.append(order)
    if not added:
        return
    if get_rate(chat_id) is None:
        try:
            await ctx.bot.send_message(
                chat_id=chat_id,
                text="⚠️ Работаешь без курса! Установи через /setrate — он применится ко всем заявкам без курса."
            )
        except Exception:
            pass
    try:
        await ctx.bot.set_message_reaction(
            chat_id=chat_id, message_id=msg.message_id,
            reaction=[ReactionTypeEmoji(emoji="👍")]
        )
    except Exception:
        pass
    lines = [f"✅ Добавлено {len(added)} заявок:"]
    for o in added:
        lines.append(f"  #{o['id']} {o['label']}")
    if skipped:
        lines.append(f"  ⚠️ Не распознано: {skipped}")
    await msg.reply_text("\n".join(lines))
    await _send_list_to_chat(ctx, chat_id)

# ─── end multi-order ──────────────────────────────────────────────


_GUARD_CARD_RE  = re.compile(r'(?<!\d)\d{4}[ \t\-]?\d{4}[ \t\-]?\d{4}[ \t\-]?\d{4}(?!\d)')
_GUARD_PHONE_RE = re.compile(r'\+?\s*[78][\s\-\(\)]*\d{3}[\s\-\(\)]*\d{3}[\s\-\(\)]*\d{2}[\s\-\(\)]*\d{2}|(?<!\d)[789]\d{9,10}(?!\d)')

_GUARD_CARD_RE  = re.compile(r'(?<!\d)\d{4}[ \t\-]?\d{4}[ \t\-]?\d{4}[ \t\-]?\d{4}(?!\d)')
_GUARD_PHONE_RE = re.compile(r'\+?\s*[78][\s\-\(\)]*\d{3}[\s\-\(\)]*\d{3}[\s\-\(\)]*\d{2}[\s\-\(\)]*\d{2}|(?<!\d)[789]\d{9,10}(?!\d)')

def _amount_looks_real(text, amount):
    """Проверяет что сумма реально есть в тексте (как отдельный токен,
    а не часть карты/телефона)."""
    if amount is None or text is None:
        return True
    try:
        amt = int(amount)
    except (TypeError, ValueError):
        return True
    masked = _GUARD_CARD_RE.sub(" ", text)
    masked = _GUARD_PHONE_RE.sub(" ", masked)
    masked_digits = re.sub(r"\D", "", masked)
    amt_str = str(amt)
    if amt_str in masked_digits:
        return True
    if amt % 1000 == 0:
        base = amt // 1000
        pat = re.compile(rf"(?<!\d){base}\s*(?:[кk]\b|т\.?\s*р|тр\b|тыс)", re.IGNORECASE)
        if pat.search(masked):
            return True
    base_frac = amt / 1000
    if base_frac != int(base_frac):
        base_s = f"{base_frac:.3f}".rstrip("0").rstrip(".")
        base_alt = base_s.replace(".", ",")
        pat = re.compile(
            rf"(?<![\d.,])(?:{re.escape(base_s)}|{re.escape(base_alt)})\s*(?:[кk]\b|т\.?\s*р|тр\b|тыс)",
            re.IGNORECASE
        )
        if pat.search(masked):
            return True
    return False

def _card_looks_real(text, card):
    """Все 16 цифр карты должны быть в тексте подряд."""
    if not card or text is None:
        return True
    card_digits = re.sub(r"\D", "", str(card))
    text_digits = re.sub(r"\D", "", text)
    return card_digits in text_digits

def _phone_looks_real(text, phone):
    """Последние 10 цифр телефона должны быть в тексте.
    Учитывает что в тексте может быть 8XXX, 7XXX, 9XXX или +7XXX — все они валидны
    если последние 10 цифр совпадают."""
    if not phone or text is None:
        return True
    phone_digits = re.sub(r"\D", "", str(phone))
    if len(phone_digits) < 10:
        return False
    last10 = phone_digits[-10:]
    text_digits = re.sub(r"\D", "", text)
    if last10 in text_digits:
        return True
    # Возможно Groq ошибся со страной (12 цифр вместо 11)
    # Тогда берём ещё короче — последние 9 цифр
    return last10[1:] in text_digits if len(last10) >= 10 else False

async def handle_message(update, ctx):
    msg = update.message
    if not msg or not msg.text:
        return
    # Частичная оплата
    if ctx.user_data.get("settle_partial_chat"):
        settle_chat = ctx.user_data.pop("settle_partial_chat")
        try:
            paid = float(msg.text.replace(",", "."))
        except ValueError:
            await msg.reply_text("❌ Введи число, например: 1500")
            ctx.user_data["settle_partial_chat"] = settle_chat
            return
        # Считаем итог в USD
        rate = get_rate(settle_chat)
        closed_orders = [o for o in get_orders(settle_chat) if o["status"] == "closed"]
        usd_total = debt.get(settle_chat, 0)
        for o in closed_orders:
            if not o.get("amount"):
                continue
            cur = o.get("currency", "RUB")
            r = o.get("close_rate") or rate
            if cur in ("USDT", "USD", "USDC"):
                usd_total += o["amount"]
            elif cur == "RUB" and r:
                usd_total += o["amount"] / r
        remaining = usd_total - paid
        # Оплаченную часть (USDT) прибавляем к балансу. Остаток уже минус в балансе с момента закрытия.
        if paid > 0:
            cur = balance.get(settle_chat)
            if cur is None:
                cur = 0
            balance[settle_chat] = cur + paid
        debt[settle_chat] = 0
        # Архивируем
        if settle_chat not in sessions:
            sessions[settle_chat] = []
        paid_rub_val = paid * rate if rate else 0
        sessions[settle_chat].append({
            "date": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "count": len(closed_orders),
            "orders": [dict(o) for o in closed_orders],
            "paid": paid,
            "debt": 0,
            "topups": list(topups.get(settle_chat, [])),
            "final_balance": balance.get(settle_chat),
            "settled_rub": paid_rub_val,
        })
        for o in get_orders(settle_chat):
            if o["status"] == "closed":
                o["status"] = "settled"
        bal_after = balance.get(settle_chat)
        bal_line = ""
        if bal_after is not None:
            if rate:
                bal_str = f"{round(bal_after * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(bal_after)
            else:
                bal_str = fmt_usd(bal_after)
            bal_line = f"\n💰 Баланс: {bal_str}"
        if paid > 0:
            if rate:
                paid_rub_str = f"{int(paid_rub_val):,}".replace(",", " ") + " ₽"
                head = f"✅ Оплачено: {fmt_usd(paid)} (+{paid_rub_str} на баланс)"
            else:
                head = f"✅ Оплачено: {fmt_usd(paid)} (+{fmt_usd(paid)} на баланс)"
        else:
            head = f"✅ Оплачено: 0 (баланс не изменился)"
        if remaining > 0:
            tail = f"\n📉 Остаток {fmt_usd(remaining)} остаётся минусом в балансе"
        else:
            tail = ""
        await msg.reply_text(
            f"{head}{tail}{bal_line}\nСписок обнулён.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]])
        )
        return

    # === Ввод произвольной суммы пополнения баланса в USDT ===
    bal_input_usdt_chat = ctx.user_data.get("balance_input_usdt_chat")
    bal_input_usdt_settle = ctx.user_data.get("balance_input_usdt_settle_chat")
    if bal_input_usdt_chat or bal_input_usdt_settle:
        target_chat = bal_input_usdt_chat or bal_input_usdt_settle
        is_settle = bool(bal_input_usdt_settle)
        s = (msg.text or "").strip().lower().replace(" ", "").replace(",", ".")
        try:
            amt_usd = float(s)
        except ValueError:
            await msg.reply_text("⚠️ Не понял сумму. Примеры: 150, 1500.5")
            return
        if amt_usd <= 0:
            await msg.reply_text("⚠️ Сумма должна быть больше нуля.")
            return
        ctx.user_data.pop("balance_input_usdt_chat", None)
        ctx.user_data.pop("balance_input_usdt_settle_chat", None)
        rate = get_rate(target_chat)
        cur = balance.get(target_chat)
        if cur is None:
            cur = 0
        balance[target_chat] = cur + amt_usd
        topups.setdefault(target_chat, []).append({
            "date": datetime.now().strftime("%H:%M"),
            "amount": int(amt_usd * rate) if rate else 0,
            "amount_usd": amt_usd,
            "rate": rate or 0,
        })
        new_bal = balance[target_chat]
        if rate:
            bal_str = f"{round(new_bal * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(new_bal)
        else:
            bal_str = fmt_usd(new_bal)
        back_cb = "balance_topup_settle" if is_settle else "balance_topup"
        keyboard_b = [
            [InlineKeyboardButton("➕ Ещё пополнить", callback_data=back_cb)],
            [InlineKeyboardButton("← В список", callback_data="go_list")],
        ]
        await msg.reply_text(
            f"✅ Баланс пополнен на {fmt_usd(amt_usd)}\n\n💰 Новый баланс: {bal_str}",
            reply_markup=InlineKeyboardMarkup(keyboard_b)
        )
        return

    # === Ввод произвольной суммы пополнения баланса ===
    bal_input_chat = ctx.user_data.get("balance_input_chat")
    bal_input_settle = ctx.user_data.get("balance_input_settle_chat")
    if bal_input_chat or bal_input_settle:
        target_chat = bal_input_chat or bal_input_settle
        is_settle = bool(bal_input_settle)
        s = (msg.text or "").strip().lower().replace(" ", "").replace(",", ".")
        m = re.match(r'^(\d+(?:\.\d+)?)([кk]|т\.?р|тр|тыс)?$', s)
        if not m:
            await msg.reply_text("⚠️ Не понял сумму. Примеры: 75000, 75к, 1.5к")
            return
        n = float(m.group(1))
        if m.group(2):
            n *= 1000
        amt = int(n)
        if amt <= 0:
            await msg.reply_text("⚠️ Сумма должна быть больше нуля.")
            return
        ctx.user_data.pop("balance_input_chat", None)
        ctx.user_data.pop("balance_input_settle_chat", None)
        rate = get_rate(target_chat)
        amt_str = f"{amt:,}".replace(",", " ") + " ₽"
        if rate:
            usdt_str = fmt_usd(amt / rate)
            rate_str = f" по курсу {rate} ₽"
        else:
            usdt_str = "— (курс не задан)"
            rate_str = ""
        if is_settle:
            confirm_cb = f"balance_add_settle:{amt}"
            back_cb = "balance_topup_settle"
        else:
            confirm_cb = f"balance_add:{amt}"
            back_cb = "balance_topup"
        keyboard_c = [
            [InlineKeyboardButton(f"✅ Подтвердить +{amt_str}", callback_data=confirm_cb)],
            [InlineKeyboardButton("← Назад", callback_data=back_cb)],
        ]
        await msg.reply_text(
            f"💰 Пополнение баланса\n\nСумма: {amt_str}\nВ USDT: ≈ {usdt_str}{rate_str}\n\nПодтверди отправку:",
            reply_markup=InlineKeyboardMarkup(keyboard_c)
        )
        return

    # Ввод банка для pending заявки
    if ctx.user_data.get("editing_bank_msg_id"):
        msg_id = ctx.user_data.pop("editing_bank_msg_id")
        parsed = ctx.user_data.pop(f"pending_{msg_id}", None)
        if parsed:
            bank_input = msg.text.strip().lower()
            bank_map = {
                "сбер": "Сбер", "сбербанк": "Сбер",
                "тинькофф": "Тинькофф", "тинк": "Тинькофф", "тинькофф": "Тинькофф",
                "втб": "ВТБ", "альфа": "Альфа", "озон": "Озон",
                "райффайзен": "Райффайзен", "газпром": "Газпромбанк",
                "открытие": "Открытие", "росбанк": "Росбанк",
                "совком": "Совкомбанк", "юмани": "ЮMoney", "qiwi": "QIWI",
                "мкб": "МКБ", "псб": "ПСБ", "уралсиб": "Уралсиб",
                "отп": "ОТП", "почта": "Почта Банк",
            }
            bank = None
            for key, name in bank_map.items():
                if key in bank_input:
                    bank = name
                    break
            if not bank:
                bank = msg.text.strip().capitalize()
            parsed["bank"] = bank
            chat_id = msg.chat_id

            # Если суммы нет — просим её указать
            if not parsed.get("amount"):
                ctx.user_data[f"pending_{msg_id}"] = parsed
                ctx.user_data["editing_amount_msg_id"] = msg_id
                await msg.reply_text(
                    f"Банк: {bank}\nТеперь укажи сумму в чат (например: 15000 или 15к):",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Игнорировать", callback_data="ignore_order")]])
                )
                return

            order_id = next_id(chat_id)
            order = {
                "id": order_id, "label": "",
                "original_text": "",
                "bank": parsed.get("bank"), "amount": parse_amount(parsed.get("amount")),
                "currency": parsed.get("currency", "RUB"),
                "card": parsed.get("card"), "phone": parsed.get("phone"),
                "rate": get_rate(chat_id), "close_rate": None,
                "status": "open", "closed_by": None, "closed_at": None,
                "message_id": msg_id,
                "added_at": datetime.now().strftime("%H:%M"),
            }
            order["label"] = format_label(order)
            get_orders(chat_id).append(order)
            try:
                await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=msg_id, reaction=[ReactionTypeEmoji(emoji="👍")])
            except Exception:
                pass
            await msg.reply_text(f"✅ Заявка #{order_id} добавлена:\n{order['label']}")
        return

    # Ввод суммы после банка
    if ctx.user_data.get("editing_amount_msg_id"):
        msg_id = ctx.user_data.pop("editing_amount_msg_id")
        parsed = ctx.user_data.pop(f"pending_{msg_id}", None)
        if parsed:
            amt = parse_amount(msg.text.strip())
            if not amt:
                await msg.reply_text("❌ Не понял сумму. Введи число, например: 15000 или 15к")
                ctx.user_data["editing_amount_msg_id"] = msg_id
                ctx.user_data[f"pending_{msg_id}"] = parsed
                return
            parsed["amount"] = amt
            chat_id = msg.chat_id
            order_id = next_id(chat_id)
            order = {
                "id": order_id, "label": "",
                "original_text": "",
                "bank": parsed.get("bank"), "amount": amt,
                "currency": parsed.get("currency", "RUB"),
                "card": parsed.get("card"), "phone": parsed.get("phone"),
                "rate": get_rate(chat_id), "close_rate": None,
                "status": "open", "closed_by": None, "closed_at": None,
                "message_id": msg_id,
                "added_at": datetime.now().strftime("%H:%M"),
            }
            order["label"] = format_label(order)
            get_orders(chat_id).append(order)
            try:
                await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=msg_id, reaction=[ReactionTypeEmoji(emoji="👍")])
            except Exception:
                pass
            await msg.reply_text(f"✅ Заявка #{order_id} добавлена:\n{order['label']}")
        return

    # Редактирование заявки
    if ctx.user_data.get("editing_order_id"):
        order_id = ctx.user_data.pop("editing_order_id")
        chat_id_edit = ctx.user_data.pop("editing_chat_id", msg.chat_id)
        missing = ctx.user_data.pop("editing_missing", [])
        order = next((o for o in get_orders(msg.chat_id) if o["id"] == order_id), None)
        if order:
            # Парсим введённый текст
            text_low = msg.text.lower()
            # Банк
            if "банк" in missing:
                for key, name in {
                    "сбер": "Сбер", "тинькофф": "Тинькофф", "тинк": "Тинькофф",
                    "втб": "ВТБ", "альфа": "Альфа", "озон": "Озон",
                    "райффайзен": "Райффайзен", "газпром": "Газпромбанк",
                    "открытие": "Открытие", "росбанк": "Росбанк",
                    "совком": "Совкомбанк", "юмани": "ЮMoney", "qiwi": "QIWI",
                    "мкб": "МКБ", "псб": "ПСБ", "уралсиб": "Уралсиб",
                }.items():
                    if key in text_low:
                        order["bank"] = name
                        break
            # Сумма
            if "сумма" in missing:
                amt = parse_amount(msg.text.strip().split()[0])
                if amt:
                    order["amount"] = amt
                    order["currency"] = "RUB"
            order["label"] = format_label(order)
            await msg.reply_text(f"✅ Заявка #{order_id} обновлена:\n{order['label']}")
        return

    if ctx.user_data.get("waiting_close_rate"):
        order_id = ctx.user_data.pop("waiting_close_rate")
        chat_id = msg.chat_id
        user = msg.from_user.first_name or msg.from_user.username or "Участник"
        try:
            rate = float(msg.text.replace(",", "."))
            await _do_close(msg, ctx, chat_id, order_id, user, close_rate=rate, show_undo=True)
        except ValueError:
            await msg.reply_text("❌ Введи число, например: 77.5\nЗаявка осталась открытой.")
        return
    if not bot_active.get(msg.chat_id, True):
        return
    # Сначала пытаемся разделить на чанки (по строкам/anchors)
    if is_multi_order(msg.text):
        chunks = split_order_chunks(msg.text)
        if len(chunks) > 1:
            await handle_multi_order(msg, ctx, chunks)
            return
    # Если в сообщении одна строка/заявка — раскрываем "N раз по X" / "N по X"
    _mult_count, _mult_base = expand_multiplier(msg.text)
    if _mult_count > 1:
        await handle_multi_order(msg, ctx, [_mult_base] * _mult_count)
        return
    parsed = await parse_with_groq(msg.text)
    if not parsed:
        return

    # Без суммы И без реквизитов — не заявка
    # Ищем карту сами — 16 цифр подряд (AI часто путает с amount)
    if not parsed.get("card"):
        card_match = re.search(r'r"\d{16}"', msg.text.replace(" ", "").replace("-", ""))
        if not card_match:
            card_match = re.search(r'\d{4}[\s\-]?\d{4}[\s\-]?\d{4}[\s\-]?\d{4}', msg.text)
        if card_match:
            digits = re.sub(r'\D', '', card_match.group())
            if len(digits) == 16:
                parsed["card"] = digits
                # Если AI записал карту как amount — обнуляем amount
                if parsed.get("amount") and len(str(int(parsed["amount"]))) >= 15:
                    parsed["amount"] = None

    # Постобработка: если card 10-11 цифр — это телефон
    if parsed.get("card") and not parsed.get("phone"):
        digits = re.sub(r"\D", "", str(parsed["card"]))
        if 10 <= len(digits) <= 11:
            parsed["phone"] = "+7" + digits[-10:]
            parsed["card"] = None

    # Ищем банк сами по тексту если AI не нашёл
    if not parsed.get("bank"):
        bank_map = {
            "тинь": "Тинькофф", "тинёк": "Тинькофф", "тинк": "Тинькофф",
            "тин ": "Тинькофф", "тинькофф": "Тинькофф", "tinkoff": "Тинькофф",
            "сбер": "Сбер", "сбербанк": "Сбер",
            "втб": "ВТБ",
            "альфа": "Альфа",
            "озон": "Озон", "ozon": "Озон",
            "райф": "Райффайзен", "райффайзен": "Райффайзен",
            "газпром": "Газпромбанк",
            "совком": "Совкомбанк",
            "отп": "ОТП",
            "мкб": "МКБ",
            "псб": "ПСБ",
            "юмани": "ЮMoney",
            "qiwi": "QIWI", "киви": "QIWI",
            "почта": "Почта Банк",
            "уралсиб": "Уралсиб",
            "хоум": "Хоум Кредит",
            "росбанк": "Росбанк",
            "открытие": "Открытие",
            "ренессанс": "Ренессанс",
            "яндекс": "Яндекс Банк", "ябанк": "Яндекс Банк",
            "абсолют": "Абсолют",
            "авангард": "Авангард",
            "акбарс": "АК Барс", "ак барс": "АК Барс",
            "бсп": "БСП",
            "вбрр": "ВБРР",
            "экспо": "Экспобанк",
            "генбанк": "Генбанк",
            "русский стандарт": "Русский Стандарт",
            "пойдём": "Пойдём", "пойдем": "Пойдём",
            "цифра": "Цифра банк",
        }
        text_lower = msg.text.lower()
        for key, name in bank_map.items():
            if key in text_lower:
                parsed["bank"] = name
                break

    # Проверяем что amount не является телефоном (10-11 цифр)
    if parsed.get("amount"):
        amt_str = str(int(parsed["amount"])) if isinstance(parsed["amount"], float) else str(parsed["amount"])
        if len(amt_str) >= 10:
            parsed["amount"] = None

    # Ищем сумму сами если нет — после карты/телефона
    if not parsed.get("amount"):
        sum_match = re.search(r'(\d+\.?\d*)\s*к', msg.text.lower())
        if sum_match:
            parsed["amount"] = float(sum_match.group(1)) * 1000
        else:
            for m in re.finditer(r'(\d{2,7})', msg.text):
                val = float(m.group(1))
                if 10 <= val <= 5_000_000:
                    parsed["amount"] = val
                    break

    # Валидация телефона — строго 10-11 цифр
    if parsed.get("phone"):
        digits = re.sub(r"\D", "", parsed["phone"])
        if len(digits) < 10 or len(digits) > 11:
            parsed["phone"] = None
        else:
            # Нормализуем в +7XXXXXXXXXX
            parsed["phone"] = "+7" + digits[-10:]

    # Валидация карты — строго 16 цифр
    if parsed.get("card"):
        digits = re.sub(r"\D", "", parsed["card"])
        if len(digits) != 16:
            parsed["card"] = None

    # Если Groq вернул кривой/отсутствующий phone — попробуем найти его сами в тексте
    _raw_text = msg.text or ""
    if not parsed.get("phone") or not _phone_looks_real(_raw_text, parsed.get("phone")):
        m = _PHONE_RE.search(_raw_text)
        if m:
            digits = re.sub(r'\D', '', m.group())
            if 10 <= len(digits) <= 11:
                parsed["phone"] = "+7" + digits[-10:]

    # Точечно отсекаем галлюцинации Groq — обнуляем поля, которых нет в тексте
    _check_text = _raw_text
    if parsed.get("amount") and not _amount_looks_real(_check_text, parsed["amount"]):
        logger.info(f"[{msg.chat_id}] Отбросил галлюцинированный amount: text={msg.text!r} amount={parsed['amount']}")
        parsed["amount"] = None
    if parsed.get("card") and not _card_looks_real(_check_text, parsed["card"]):
        logger.info(f"[{msg.chat_id}] Отбросил галлюцинированную card: text={msg.text!r} card={parsed['card']}")
        parsed["card"] = None
    if parsed.get("phone") and not _phone_looks_real(_check_text, parsed["phone"]):
        logger.info(f"[{msg.chat_id}] Отбросил галлюцинированный phone: text={msg.text!r} phone={parsed['phone']}")
        parsed["phone"] = None
    # Жёсткая проверка формата: только +7 9XXXXXXXXX (мобильные, 11 цифр)
    if parsed.get("phone") and not re.match(r'^\+7\d{10}$', parsed["phone"]):
        logger.info(f"[{msg.chat_id}] Отбросил кривой phone: {parsed['phone']!r}")
        parsed["phone"] = None

    if not parsed.get("amount") and not parsed.get("card") and not parsed.get("phone"):
        return

    # Реквизиты обязательны — без карты/телефона не считаем за заявку
    if not parsed.get("card") and not parsed.get("phone"):
        logger.info(f"[{msg.chat_id}] Отброшено: нет реквизитов (text={msg.text!r})")
        return


    # Есть реквизиты без суммы — предупреждаем
    if not parsed.get("amount") and (parsed.get("card") or parsed.get("phone")):
        chat_id = msg.chat_id
        if not bot_active.get(chat_id, True):
            return
        req = f"···{parsed['card'][-4:]}" if parsed.get("card") else f"···{parsed['phone'][-4:]}"
        sent = await msg.reply_text(
            f"⚠️ Получены реквизиты {req} без суммы.\n\nДобавить в список или игнорировать?",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("➕ Добавить без суммы", callback_data=f"force_add:{msg.message_id}"),
                InlineKeyboardButton("❌ Игнорировать", callback_data="ignore_order"),
            ]])
        )
        # Сохраняем parsed для возможного добавления
        ctx.user_data[f"pending_{msg.message_id}"] = parsed
        return

    chat_id = msg.chat_id
    # Защита: банк обязателен. Без банка — спрашиваем.
    if not parsed.get("bank"):
        if not bot_active.get(chat_id, True):
            return
        if parsed.get("card"):
            req = f"···{parsed['card'][-4:]}"
        elif parsed.get("phone"):
            req = f"···{parsed['phone'][-4:]}"
        else:
            req = ""
        amt_str = ""
        if parsed.get("amount"):
            amt_str = f" · {int(parsed['amount']):,} {parsed.get('currency','RUB')}".replace(",", " ")
        await msg.reply_text(
            f"⚠️ Банк не распознан ({req}{amt_str}).\n\nВыбери банк или игнорируй:",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("✏️ Выбрать банк", callback_data=f"edit_bank:{msg.message_id}"),
                InlineKeyboardButton("❌ Игнорировать", callback_data="ignore_order"),
            ]])
        )
        ctx.user_data[f"pending_{msg.message_id}"] = parsed
        return
    order_id = next_id(chat_id)
    order = {
        "id": order_id, "label": "",
        "original_text": msg.text,
        "bank": parsed.get("bank"), "amount": parsed.get("amount"),
        "currency": parsed.get("currency", "RUB"),
        "card": parsed.get("card"), "phone": parsed.get("phone"),
        "rate": get_rate(chat_id), "close_rate": None,
        "status": "open", "closed_by": None, "closed_at": None,
        "message_id": msg.message_id,
        "added_at": datetime.now().strftime("%H:%M"),
    }
    order["label"] = format_label(order)
    get_orders(chat_id).append(order)
    try:
        await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=msg.message_id, reaction=[ReactionTypeEmoji(emoji="👍")])
    except Exception as e:
        logger.warning(f"Реакция: {e}")
    logger.info(f"[{chat_id}] Заявка #{order_id}: {order['label']}")
    # Предупреждение если работаем без курса
    if get_rate(chat_id) is None:
        try:
            await ctx.bot.send_message(
                chat_id=chat_id,
                text="⚠️ Работаешь без курса! Установи через /setrate — он применится ко всем заявкам без курса."
            )
        except Exception:
            pass
    await _send_list_to_chat(ctx, chat_id)

async def handle_delete(update, ctx):
    msg = update.message
    if not msg:
        return
    for o in get_orders(msg.chat_id):
        if o["message_id"] == msg.message_id and o["status"] == "open":
            o["status"] = "deleted"
            break

async def show_list(update, ctx):
    chat_id = update.effective_chat.id
    logger.info(f"show_list called, chat_id={chat_id}, orders={get_orders(chat_id)}")
    open_orders = [o for o in get_orders(chat_id) if o["status"] == "open"]
    closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
    settled_orders = [o for o in get_orders(chat_id) if o["status"] == "settled"]
    rate_text = f"💱 Курс: {current_rate[chat_id]} ₽\n" if chat_id in current_rate else ""
    has_debt = debt.get(chat_id, 0) > 0
    has_sessions = bool(sessions.get(chat_id))

    # Удаляем сообщение пользователя с /list чтобы не засирать чат
    try:
        await ctx.bot.delete_message(chat_id=chat_id, message_id=update.message.message_id)
    except Exception:
        pass

    if not open_orders and not closed_orders and not settled_orders and not has_debt and not has_sessions:
        text = f"{rate_text}{_balance_line(chat_id)}📭 Активных заявок нет. Отправляй заявки в чат."
        keyboard = [[InlineKeyboardButton("💰 Пополнить баланс", callback_data="balance_topup")]]
    else:
        keyboard = [[InlineKeyboardButton(f"#{o['id']}  {o['label']}  ({o['added_at']})", callback_data=f"ask:{o['id']}")] for o in open_orders]
        btn_row = []
        if closed_orders or settled_orders:
            btn_row.append(InlineKeyboardButton("💰 Результат дня", callback_data="show_total"))
        if closed_orders:
            btn_row.append(InlineKeyboardButton("🧾 Посчитаться", callback_data="calculate"))
        if btn_row:
            keyboard.append(btn_row)
        if sessions.get(chat_id):
            keyboard.append([InlineKeyboardButton("📒 Логи сессий", callback_data="show_logs")])
        last_closed = next((o for o in reversed(get_orders(chat_id)) if o["status"] == "closed"), None)
        if last_closed:
            keyboard.append([InlineKeyboardButton(f"↩️ Отменить закрытие #{last_closed['id']}", callback_data=f"undo:{last_closed['id']}")])
        _bulk = last_bulk_close.get(chat_id) or []
        _still = [oid for oid in _bulk if any(o["id"]==oid and o["status"]=="closed" for o in get_orders(chat_id))]
        if len(_still) >= 2:
            keyboard.append([InlineKeyboardButton(f"↩️ Отменить закрытие всех ({len(_still)})", callback_data="undo_all")])
        keyboard.append([InlineKeyboardButton("💰 Пополнить баланс", callback_data="balance_topup")])
        if open_orders:
            keyboard.append([InlineKeyboardButton(f"✅ Закрыть все ({len(open_orders)})", callback_data="close_all")])
            keyboard.append([InlineKeyboardButton("🗑 Удалить заявку", callback_data="del_menu")])
        text = rate_text + _balance_line(chat_id) + (f"📋 Активные заявки — {len(open_orders)} шт.\nНажми чтобы закрыть:" if open_orders else "📭 Активных заявок нет.")

    markup = InlineKeyboardMarkup(keyboard) if keyboard else None

    # Удаляем предыдущее /list-меню (если было) и шлём новое
    prev_id = last_list_msg.get(chat_id)
    if prev_id:
        try:
            await ctx.bot.delete_message(chat_id=chat_id, message_id=prev_id)
        except Exception:
            pass
        # Также убираем из окна авто-очистки чтобы не пытаться удалить ещё раз
        try:
            if chat_id in bot_msgs and prev_id in bot_msgs[chat_id]:
                bot_msgs[chat_id].remove(prev_id)
        except Exception:
            pass
    sent = await ctx.bot.send_message(chat_id=chat_id, text=text, reply_markup=markup)
    last_list_msg[chat_id] = sent.message_id

async def _do_close(msg_or_query, ctx, chat_id, order_id, user, close_rate, show_undo=False):
    order = next((o for o in get_orders(chat_id) if o["id"] == order_id), None)
    if not order or order["status"] != "open":
        return
    order["status"] = "closed"
    order["closed_by"] = user
    order["closed_at"] = datetime.now().strftime("%H:%M")
    order["close_rate"] = close_rate
    try:
        deduct_usd = _order_to_usd(order, get_rate(chat_id))
        if deduct_usd:
            cur_bal = balance.get(chat_id)
            if cur_bal is None:
                cur_bal = 0
            balance[chat_id] = cur_bal - deduct_usd
    except Exception:
        pass
    try:
        await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=order["message_id"], reaction=[ReactionTypeEmoji(emoji="🎉")])
    except Exception:
        pass
    rate_str = f"\nКурс сделки: {close_rate}" if close_rate else ""
    text = f"✅ Заявка #{order_id} закрыта\n{order['label']}{rate_str}\nЗакрыл: {user} · {order['closed_at']}"
    keyboard = [[InlineKeyboardButton("Отменить", callback_data=f"undo:{order_id}"), InlineKeyboardButton("В список", callback_data="go_list")], [InlineKeyboardButton("🧾 Посчитаться", callback_data="calculate")]] if show_undo else None
    markup = InlineKeyboardMarkup(keyboard) if keyboard else None
    if hasattr(msg_or_query, "edit_message_text"):
        await msg_or_query.edit_message_text(text, reply_markup=markup)
    else:
        await msg_or_query.reply_text(text, reply_markup=markup)


async def _show_list_inline(query, ctx, chat_id):
    open_orders   = [o for o in get_orders(chat_id) if o["status"] == "open"]
    closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
    settled_orders = [o for o in get_orders(chat_id) if o["status"] == "settled"]
    rate_text = f"💱 Курс: {current_rate[chat_id]} ₽\n" if chat_id in current_rate else ""
    keyboard = [[InlineKeyboardButton(f"#{o['id']}  {o['label']}  ({o['added_at']})", callback_data=f"ask:{o['id']}")] for o in open_orders]
    btn_row = []
    if closed_orders or settled_orders:
        btn_row.append(InlineKeyboardButton("💰 Результат дня", callback_data="show_total"))
    if closed_orders:
        btn_row.append(InlineKeyboardButton("🧾 Посчитаться", callback_data="calculate"))
    if btn_row:
        keyboard.append(btn_row)
    if sessions.get(chat_id):
        keyboard.append([InlineKeyboardButton("📒 Логи сессий", callback_data="show_logs")])
    last_closed = next((o for o in reversed(get_orders(chat_id)) if o["status"] == "closed"), None)
    if last_closed:
        keyboard.append([InlineKeyboardButton(f"Отменить закрытие #{last_closed['id']}", callback_data=f"undo:{last_closed['id']}")])
    _bulk = last_bulk_close.get(chat_id) or []
    _still = [oid for oid in _bulk if any(o["id"]==oid and o["status"]=="closed" for o in get_orders(chat_id))]
    if len(_still) >= 2:
        keyboard.append([InlineKeyboardButton(f"↩️ Отменить закрытие всех ({len(_still)})", callback_data="undo_all")])
    keyboard.append([InlineKeyboardButton("💰 Пополнить баланс", callback_data="balance_topup")])
    if open_orders:
        keyboard.append([InlineKeyboardButton(f"✅ Закрыть все ({len(open_orders)})", callback_data="close_all")])
        keyboard.append([InlineKeyboardButton("🗑 Удалить заявку", callback_data="del_menu")])
    text = rate_text + _balance_line(chat_id) + (f"📋 Активные заявки — {len(open_orders)} шт.\nНажми чтобы закрыть:" if open_orders else "📭 Активных заявок нет.")
    try:
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    except Exception as e:
        if chat_id in last_bot_msg:
            try:
                await ctx.bot.delete_message(chat_id=chat_id, message_id=last_bot_msg[chat_id])
            except Exception:
                pass
        msg = await ctx.bot.send_message(chat_id=chat_id, text=text, reply_markup=InlineKeyboardMarkup(keyboard))
        last_bot_msg[chat_id] = msg.message_id

async def _try_delete_source_message(ctx, chat_id, message_id, all_orders):
    """Удаляет исходное сообщение в чате, если оно содержало только эту одну заявку.
    Возвращает True если сообщение удалено."""
    siblings = [o for o in all_orders if o["message_id"] == message_id]
    if len(siblings) != 1:
        return False
    try:
        await ctx.bot.delete_message(chat_id=chat_id, message_id=message_id)
        return True
    except Exception:
        return False

def build_excel_report(chat_id):
    """Генерирует Excel-отчёт со всеми данными по chat_id. Возвращает BytesIO."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    rate = get_rate(chat_id)
    orders_list = [o for o in get_orders(chat_id) if o["status"] in ("closed", "settled")]
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["#", "Время", "Банк", "Сумма", "Валюта", "Курс", "USD", "Карта", "Телефон", "Статус"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    usd_total = 0
    rub_total = 0
    for o in orders_list:
        amt = o.get("amount") or 0
        cur = o.get("currency", "RUB")
        r = o.get("close_rate") or o.get("rate") or rate
        if cur in ("USDT", "USD", "USDC"):
            order_usd = amt
        elif cur == "RUB" and r:
            order_usd = amt / r
            rub_total += amt
        else:
            order_usd = 0
            if cur == "RUB":
                rub_total += amt
        usd_total += order_usd
        ws.append([
            o.get("id"),
            o.get("closed_at") or o.get("added_at") or "",
            o.get("bank") or "?",
            amt,
            cur,
            r or "",
            round(order_usd, 3) if order_usd else "",
            str(o.get("card") or ""),
            str(o.get("phone") or ""),
            o.get("status"),
        ])
        for cell in ws[ws.max_row]:
            cell.border = border
    ws.append([])
    ws.append(["", "", "Итого RUB:", rub_total, "", "", "", "", "", ""])
    ws[ws.max_row][2].font = bold
    ws[ws.max_row][3].font = bold
    ws.append(["", "", "Итого USD:", "", "", "", round(usd_total, 3), "", "", ""])
    ws[ws.max_row][2].font = bold
    ws[ws.max_row][6].font = bold
    ws.append([])
    # Пополнения
    tps = topups.get(chat_id) or []
    if tps:
        ws.append(["💰 Пополнения баланса:"])
        ws[ws.max_row][0].font = bold
        ws.append(["Время", "RUB", "USD", "Курс"])
        for cell in ws[ws.max_row]:
            cell.font = bold
        topup_rub_sum = 0
        topup_usd_sum = 0
        for t in tps:
            a_rub = t.get("amount") or 0
            a_usd = t.get("amount_usd") or 0
            tr = t.get("rate") or ""
            topup_rub_sum += a_rub
            topup_usd_sum += a_usd
            ws.append([t.get("date", ""), a_rub, round(a_usd, 3) if a_usd else "", tr])
        ws.append(["Итого:", topup_rub_sum, round(topup_usd_sum, 3) if topup_usd_sum else "", ""])
        ws[ws.max_row][0].font = bold
        ws[ws.max_row][1].font = bold
        ws[ws.max_row][2].font = bold
        ws.append([])
    # Итоговый баланс
    bal_usd = balance.get(chat_id)
    ws.append(["💼 Текущий баланс:"])
    ws[ws.max_row][0].font = bold
    if bal_usd is None:
        ws.append(["Не задан"])
    else:
        bal_rub = bal_usd * rate if rate else None
        ws.append(["В USD:", round(bal_usd, 3)])
        ws[ws.max_row][0].font = bold
        if bal_rub is not None:
            ws.append(["В RUB:", int(bal_rub)])
            ws[ws.max_row][0].font = bold
        if rate:
            ws.append(["По курсу:", rate])
            ws[ws.max_row][0].font = bold
    widths = [6, 10, 14, 14, 10, 8, 14, 20, 18, 10]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = w
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

async def handle_callback(update, ctx):
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat_id
    user = query.from_user.first_name or query.from_user.username or "Участник"
    data = query.data

    if data.startswith("ask:"):
        order_id = int(data.split(":")[1])
        order = next((o for o in get_orders(chat_id) if o["id"] == order_id), None)
        if not order or order["status"] != "open":
            await query.edit_message_text("⚠️ Заявка уже закрыта.")
            return
        keyboard = [
            [InlineKeyboardButton("✅ Закрыть", callback_data=f"confirm:{order_id}")],
            [InlineKeyboardButton("💱 Закрыть с курсом", callback_data=f"enter_rate_pre:{order_id}")],
            [InlineKeyboardButton("← Меню", callback_data="back_list")]
        ]
        await query.edit_message_text(f"Закрыть заявку #{order['id']}?\n\n{order['label']}", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("confirm:"):
        order_id = int(data.split(":")[1])
        order = next((o for o in get_orders(chat_id) if o["id"] == order_id), None)
        if not order or order["status"] != "open":
            await query.edit_message_text("⚠️ Заявка уже закрыта.")
            return
        await _do_close(query, ctx, chat_id, order_id, user, close_rate=None, show_undo=True)

    elif data.startswith("enter_rate_pre:"):
        order_id = int(data.split(":")[1])
        ctx.user_data["waiting_close_rate"] = order_id
        keyboard = [[InlineKeyboardButton("← Меню", callback_data=f"ask:{order_id}")]]
        await query.edit_message_text(f"Введи курс для заявки #{order_id} в чат:\n(после ввода заявка закроется)", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("enter_rate:"):
        order_id = int(data.split(":")[1])
        ctx.user_data["waiting_close_rate"] = order_id
        keyboard = [[InlineKeyboardButton("← Меню", callback_data=f"confirm:{order_id}")]]
        await query.edit_message_text(f"Введи курс сделки для заявки #{order_id} в чат:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("close_no_rate:"):
        order_id = int(data.split(":")[1])
        ctx.user_data.pop("waiting_close_rate", None)
        await _do_close(query, ctx, chat_id, order_id, user, close_rate=None)

    elif data == "show_total":
        closed_orders = [o for o in get_orders(chat_id) if o["status"] in ("closed", "settled")]
        if not closed_orders:
            await query.answer("Нет исполненных заявок", show_alert=True)
            return

        rate = get_rate(chat_id)
        order_lines = []
        totals = {}
        usd_total = 0

        for o in closed_orders:
            amt = o.get("amount")
            if amt is None:
                continue
            cur = o.get("currency", "RUB")
            totals[cur] = totals.get(cur, 0) + amt

            # Курс: close_rate (если указан при закрытии), иначе rate заявки, иначе текущий курс чата
            r = o.get("close_rate") or o.get("rate") or rate

            # Реквизиты (последние 4 цифры)
            req = ""
            if o.get("card"):
                req = f" · ···{o['card'][-4:]}"
            elif o.get("phone"):
                req = f" · ···{o['phone'][-4:]}"

            amt_str = f"{int(amt):,}".replace(",", " ") if float(amt).is_integer() else str(amt)
            bank = o.get("bank") or "?"

            # USD для этой заявки и общего итога
            line_usd = ""
            if cur in ("USDT", "USD", "USDC"):
                usd_total += amt
            elif cur == "RUB" and r:
                order_usd = amt / r
                usd_total += order_usd
                line_usd = f" → {fmt_usd(order_usd)}"

            rate_str = f" · курс {r}" if r and cur == "RUB" else ""
            order_lines.append(f"#{o['id']} {bank} · {amt_str} {cur}{rate_str}{req}{line_usd}")

        totals_lines = [f"  {int(v):,} {k}".replace(",", " ") for k, v in totals.items()]

        lines = [f"📊 Исполнено заявок: {len(closed_orders)}", ""] + order_lines + ["", "💰 По валютам:"] + totals_lines

        if usd_total > 0:
            lines.append("")
            lines.append(f"💵 Итого ≈ {fmt_usd(usd_total)}")
        elif not rate:
            lines.append("")
            lines.append("💡 Установи курс (/setrate) для подсчёта в USD")

        text = "\n".join(lines)
        if len(text) > 4000:
            text = text[:3990] + "\n..."

        await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📥 Скачать в Excel", callback_data="report_excel")],
                [InlineKeyboardButton("← В список", callback_data="go_list")],
            ])
        )

    elif data == "report_excel":
        try:
            from telegram import InputFile
            ord_list = [o for o in get_orders(chat_id) if o["status"] in ("closed", "settled")]
            if not ord_list and not topups.get(chat_id) and balance.get(chat_id) is None:
                await query.answer("Нет данных для отчёта", show_alert=True)
                return
            bio = build_excel_report(chat_id)
            fname = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            await ctx.bot.send_document(
                chat_id=chat_id,
                document=InputFile(bio, filename=fname),
                caption="📊 Отчёт за день"
            )
        except Exception as e:
            logger.error(f"Excel error: {e}")
            try:
                await ctx.bot.send_message(chat_id=chat_id, text=f"❌ Ошибка генерации Excel: {e}")
            except Exception:
                pass

    elif data == "undo_all":
        bulk = last_bulk_close.get(chat_id) or []
        restored = 0
        for oid in bulk:
            order = next((o for o in get_orders(chat_id) if o["id"] == oid), None)
            if not order or order["status"] != "closed":
                continue
            try:
                add_back_usd = _order_to_usd(order, get_rate(chat_id))
                if add_back_usd:
                    cur_bal = balance.get(chat_id)
                    if cur_bal is None:
                        cur_bal = 0
                    balance[chat_id] = cur_bal + add_back_usd
            except Exception:
                pass
            order["status"] = "open"
            order["closed_by"] = None
            order["closed_at"] = None
            order["close_rate"] = None
            try:
                await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=order["message_id"], reaction=[ReactionTypeEmoji(emoji="👍")])
            except Exception:
                pass
            restored += 1
        last_bulk_close.pop(chat_id, None)
        await query.answer(f"Восстановлено заявок: {restored}", show_alert=False)
        await _show_list_inline(query, ctx, chat_id)

    elif data.startswith("undo:"):
        order_id = int(data.split(":")[1])
        order = next((o for o in get_orders(chat_id) if o["id"] == order_id), None)
        if not order or order["status"] != "closed":
            await query.edit_message_text("⚠️ Нельзя отменить.")
            return
        try:
            add_back_usd = _order_to_usd(order, get_rate(chat_id))
            if add_back_usd:
                cur_bal = balance.get(chat_id)
                if cur_bal is None:
                    cur_bal = 0
                balance[chat_id] = cur_bal + add_back_usd
        except Exception:
            pass
        order["status"] = "open"
        order["closed_by"] = None
        order["closed_at"] = None
        order["close_rate"] = None
        try:
            await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=order["message_id"], reaction=[ReactionTypeEmoji(emoji="👍")])
        except Exception:
            pass
        await query.edit_message_text(f"↩️ Закрытие заявки #{order_id} отменено.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В список", callback_data="go_list")]]))

    elif data == "calculate":
        closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
        if not closed_orders:
            await query.answer("Нет выполненных заявок", show_alert=True)
            return
        rate = get_rate(chat_id)
        totals = {}
        usd_total = 0
        for o in closed_orders:
            if not o.get("amount"):
                continue
            cur = o.get("currency", "RUB")
            totals[cur] = totals.get(cur, 0) + o["amount"]
            r = o.get("close_rate") or rate
            if cur in ("USDT", "USD", "USDC"):
                usd_total += o["amount"]
            elif cur == "RUB" and r:
                usd_total += o["amount"] / r
        sum_lines = [f"  {int(v):,} {k}".replace(",", " ") for k, v in totals.items()]
        usd_line = f"\n💵 Итого ≈ {fmt_usd(usd_total)}" if usd_total > 0 else ("\n💡 Установи курс (/setrate) для USD" if not rate else "")
        debt_text = f"\n\n⚠️ Долг с прошлой сессии: {fmt_usd(debt.get(chat_id, 0))}" if debt.get(chat_id) else ""
        bal_text = ""
        cur_bal = balance.get(chat_id)
        if cur_bal is not None:
            if rate:
                bal_str = f"{round(cur_bal * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(cur_bal)
            else:
                bal_str = fmt_usd(cur_bal)
            bal_text = f"\n\n💰 Баланс: {bal_str}"
            if cur_bal < 0:
                need_usd = -cur_bal
                need_str = fmt_usd(need_usd)
                if rate:
                    need_str = f"{round(need_usd * rate):,}".replace(",", " ") + " ₽ / " + need_str
                bal_text += f"\n📉 Не хватает: {need_str}"
            elif cur_bal > 0:
                bal_text += f"\n✅ Остаток в плюсе"
        text = (
            f"🧾 Итог сессии ({len(closed_orders)} заявок):\n\n"
            + "\n".join(sum_lines)
            + usd_line
            + debt_text
            + bal_text
            + "\n\n📊 Зачислить за заявки — прибавить сумму всех закрытых заявок к балансу."
            "\n💸 Частично — ввести сумму оплаты, она зачислится на баланс."
            "\n🔄 Обнулить баланс — довести текущий баланс до 0 (независимо от заявок)."
            "\n💰 Пополнить баланс — добавить РУБ в кассу вручную."
        )
        full_lbl = "📊 Зачислить за заявки"
        if usd_total > 0:
            full_lbl += f" +{fmt_usd(usd_total)}"
        bal_lbl = "💰 Пополнить баланс"
        cur_bal_v = balance.get(chat_id)
        if cur_bal_v is not None:
            if rate:
                bal_lbl += f" (сейчас {round(cur_bal_v * rate):,} ₽)".replace(",", " ")
            else:
                bal_lbl += f" (сейчас {fmt_usd(cur_bal_v)})"
        # Кнопка «Обнулить баланс» — показываем только если баланс не нулевой
        rows = [
            [InlineKeyboardButton(full_lbl, callback_data="settle_confirm"),
             InlineKeyboardButton("💸 Частично", callback_data="settle_partial")],
        ]
        if cur_bal_v is not None and cur_bal_v < 0:
            zero_lbl = "🔄 Обнулить баланс"
            delta_usd = -cur_bal_v
            sign = "+" if delta_usd > 0 else "−"
            zero_lbl += f" ({sign}{fmt_usd(abs(delta_usd))})"
            rows.append([InlineKeyboardButton(zero_lbl, callback_data="zero_confirm")])
        rows.append([InlineKeyboardButton(bal_lbl, callback_data="balance_topup_settle")])
        rows.append([InlineKeyboardButton("← В список", callback_data="go_list")])
        keyboard_calc = rows
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard_calc))

    elif data == "settle_partial":
        ctx.user_data["settle_partial_chat"] = chat_id
        keyboard_p = [[InlineKeyboardButton("← Назад", callback_data="calculate")]]
        await query.edit_message_text(
            "Введи сколько оплачено в USD:",
            reply_markup=InlineKeyboardMarkup(keyboard_p)
        )

    elif data == "zero_confirm":
        cur_bal = balance.get(chat_id)
        if cur_bal is None or abs(cur_bal) < 0.0001:
            await query.answer("Баланс уже 0", show_alert=True)
            return
        rate = get_rate(chat_id)
        bal_str = fmt_usd(cur_bal)
        if rate:
            bal_str = f"{round(cur_bal * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(cur_bal)
        delta_usd = -cur_bal
        sign = "+" if delta_usd > 0 else "−"
        delta_str = f"{sign}{fmt_usd(abs(delta_usd))}"
        closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
        arch_note = f"\n📁 Архивируется {len(closed_orders)} закрытых заявок" if closed_orders else ""
        await query.edit_message_text(
            f"🔄 Обнулить баланс?\n\nТекущий баланс: {bal_str}\nКорректировка: {delta_str}{arch_note}\n\nПодтверди:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Точно обнулить", callback_data="zero")],
                [InlineKeyboardButton("← Отмена", callback_data="calculate")],
            ])
        )

    elif data == "zero":
        cur_bal = balance.get(chat_id)
        if cur_bal is None or abs(cur_bal) < 0.0001:
            await query.answer("Баланс уже 0", show_alert=True)
            return
        closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
        rate = get_rate(chat_id)
        delta_usd = -cur_bal  # корректировка чтобы стал 0
        # Архивируем сессию (даже если closed_orders пуст — фиксируем операцию обнуления)
        if chat_id not in sessions:
            sessions[chat_id] = []
        sessions[chat_id].append({
            "date": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "count": len(closed_orders),
            "orders": [dict(o) for o in closed_orders],
            "paid": cur_bal if cur_bal > 0 else None,
            "debt": -cur_bal if cur_bal < 0 else 0,
            "topups": list(topups.get(chat_id, [])),
            "final_balance": 0.0,
            "zeroed_from": cur_bal,
        })
        for o in get_orders(chat_id):
            if o["status"] == "closed":
                o["status"] = "settled"
        balance[chat_id] = 0.0
        debt[chat_id] = 0
        topups.pop(chat_id, None)
        last_bulk_close.pop(chat_id, None)
        sign = "+" if delta_usd > 0 else "−"
        op_str = f"{sign}{fmt_usd(abs(delta_usd))}"
        msg_lines = [f"🔄 Баланс обнулён. Корректировка: {op_str}"]
        if closed_orders:
            msg_lines.append(f"📁 Архивировано заявок: {len(closed_orders)}")
        msg_lines.append("💰 Баланс: 0 USD")
        await query.edit_message_text(
            "\n".join(msg_lines),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]])
        )

    elif data == "settle_confirm":
        closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
        cur_bal_c = balance.get(chat_id)
        if not closed_orders and not (cur_bal_c is not None and cur_bal_c < 0):
            await query.answer("Нет данных для закрытия", show_alert=True)
            return
        rate = get_rate(chat_id)
        sum_rub_preview = 0
        for o in closed_orders:
            amt = o.get("amount") or 0
            cur_o = o.get("currency", "RUB")
            r = o.get("close_rate") or rate
            if cur_o == "RUB":
                sum_rub_preview += amt
            elif cur_o in ("USDT", "USD", "USDC") and r:
                sum_rub_preview += amt * r
        if not closed_orders and cur_bal_c is not None and cur_bal_c < 0:
            sum_rub_preview = -cur_bal_c
        sum_str = f"{int(sum_rub_preview):,}".replace(",", " ") + " ₽"
        usd_part = f" / ≈ {fmt_usd(sum_rub_preview / rate)}" if rate else ""
        n = len(closed_orders) if closed_orders else 0
        head = (f"Закрыть сессию из {n} заявок?" if n else "Закрыть минусовой баланс?")
        await query.edit_message_text(
            f"⚠️ {head}\n\nЗачислится на баланс: +{sum_str}{usd_part}\n\nПодтверди:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Точно посчитаться", callback_data="settle")],
                [InlineKeyboardButton("← Отмена", callback_data="calculate")],
            ])
        )

    elif data == "settle":
        closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
        if not closed_orders:
            await query.edit_message_text("⚠️ Нет выполненных заявок.")
            return
        # Сумма заявок в RUB → прибавить к балансу (контрагент оплатил всё)
        rate = get_rate(chat_id)
        sum_rub = 0
        usd_paid_total = 0
        for o in closed_orders:
            amt = o.get("amount") or 0
            cur = o.get("currency", "RUB")
            r = o.get("close_rate") or rate
            if cur == "RUB":
                sum_rub += amt
                if r:
                    usd_paid_total += amt / r
            elif cur in ("USDT", "USD", "USDC"):
                usd_paid_total += amt
                if r:
                    sum_rub += amt * r
        if usd_paid_total > 0:
            cur_bal = balance.get(chat_id)
            if cur_bal is None:
                cur_bal = 0
            balance[chat_id] = cur_bal + usd_paid_total
        # Архивируем сессию
        if chat_id not in sessions:
            sessions[chat_id] = []
        sessions[chat_id].append({
            "date": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "count": len(closed_orders),
            "orders": [dict(o) for o in closed_orders],
            "paid": usd_paid_total,
            "debt": 0,
            "topups": list(topups.get(chat_id, [])),
            "final_balance": balance.get(chat_id),
            "settled_rub": sum_rub,
        })
        # Закрытые становятся settled
        for o in get_orders(chat_id):
            if o["status"] == "closed":
                o["status"] = "settled"
        debt[chat_id] = 0
        bal_after = balance.get(chat_id)
        bal_line = ""
        if bal_after is not None:
            bal_str = f"{int(bal_after):,}".replace(",", " ") + " ₽"
            if rate:
                bal_str += f" / {fmt_usd(bal_after / rate)}"
            bal_line = f"\n💰 Баланс: {bal_str}"
        sum_str = f"{int(sum_rub):,}".replace(",", " ") + " ₽"
        await query.edit_message_text(
            f"✅ Посчитались! {len(closed_orders)} заявок.\n"
            f"💸 Зачислено на баланс: +{sum_str}{bal_line}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]])
        )

    elif data == "close_all":
        open_orders = [o for o in get_orders(chat_id) if o["status"] == "open"]
        if not open_orders:
            await query.answer("Нет открытых заявок", show_alert=True)
            return
        rate = get_rate(chat_id)
        sum_rub = 0
        for o in open_orders:
            amt = o.get("amount") or 0
            if (o.get("currency", "RUB")) == "RUB":
                sum_rub += amt
            elif rate:
                sum_rub += amt * rate
        sum_str = f"{int(sum_rub):,}".replace(",", " ") + " ₽"
        usd_part = f" / ≈ {fmt_usd(sum_rub / rate)}" if rate else ""
        await query.edit_message_text(
            f"⚠️ Закрыть ВСЕ открытые заявки?\n\nЗаявок: {len(open_orders)}\nСумма: {sum_str}{usd_part}\n\n(списание с баланса по текущему курсу)",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Точно закрыть", callback_data="close_all_confirm")],
                [InlineKeyboardButton("← Отмена", callback_data="go_list")],
            ])
        )

    elif data == "close_all_confirm":
        open_orders = [o for o in get_orders(chat_id) if o["status"] == "open"]
        if not open_orders:
            await query.answer("Нет открытых заявок", show_alert=True)
            return
        closed_n = 0
        for order in open_orders:
            order["status"] = "closed"
            order["closed_by"] = user
            order["closed_at"] = datetime.now().strftime("%H:%M")
            order["close_rate"] = None
            try:
                deduct_usd = _order_to_usd(order, get_rate(chat_id))
                if deduct_usd:
                    cur_bal = balance.get(chat_id)
                    if cur_bal is None:
                        cur_bal = 0
                    balance[chat_id] = cur_bal - deduct_usd
            except Exception:
                pass
            try:
                await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=order["message_id"], reaction=[ReactionTypeEmoji(emoji="🎉")])
            except Exception:
                pass
            closed_n += 1
        last_bulk_close[chat_id] = [o["id"] for o in open_orders]
        await query.answer(f"Закрыто заявок: {closed_n}", show_alert=False)
        await _show_list_inline(query, ctx, chat_id)

    elif data == "balance_custom_usdt":
        ctx.user_data["balance_input_usdt_chat"] = chat_id
        for k in ("balance_input_chat", "balance_input_settle_chat", "balance_input_usdt_settle_chat"):
            ctx.user_data.pop(k, None)
        await query.edit_message_text(
            "💵 Введи сумму пополнения в USDT\n(например: 150, 1500.5)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← Назад", callback_data="balance_topup")]])
        )

    elif data == "balance_custom_usdt_settle":
        ctx.user_data["balance_input_usdt_settle_chat"] = chat_id
        for k in ("balance_input_chat", "balance_input_settle_chat", "balance_input_usdt_chat"):
            ctx.user_data.pop(k, None)
        await query.edit_message_text(
            "💵 Введи сумму пополнения в USDT\n(например: 150, 1500.5)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← Назад", callback_data="balance_topup_settle")]])
        )

    elif data == "balance_custom":
        ctx.user_data["balance_input_chat"] = chat_id
        ctx.user_data.pop("balance_input_settle_chat", None)
        await query.edit_message_text(
            "✏️ Введи сумму пополнения в чат\n(например: 75000, 75к, 1.5к)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← Назад", callback_data="balance_topup")]])
        )

    elif data == "balance_custom_settle":
        ctx.user_data["balance_input_settle_chat"] = chat_id
        ctx.user_data.pop("balance_input_chat", None)
        await query.edit_message_text(
            "✏️ Введи сумму пополнения в чат\n(например: 75000, 75к, 1.5к)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← Назад", callback_data="balance_topup_settle")]])
        )

    elif data.startswith("balance_confirm_settle:"):
        try:
            amt = int(data.split(":")[1])
            rate = get_rate(chat_id)
            amt_str = f"{amt:,}".replace(",", " ") + " ₽"
            if rate:
                usdt_str = fmt_usd(amt / rate)
                rate_str = f" по курсу {rate} ₽"
            else:
                usdt_str = "— (курс не задан)"
                rate_str = ""
            keyboard_c = [
                [InlineKeyboardButton(f"✅ Подтвердить +{amt_str}", callback_data=f"balance_add_settle:{amt}")],
                [InlineKeyboardButton("← Назад", callback_data="balance_topup_settle")],
            ]
            await query.edit_message_text(
                f"💰 Пополнение баланса\n\nСумма: {amt_str}\nВ USDT: ≈ {usdt_str}{rate_str}\n\nПодтверди отправку:",
                reply_markup=InlineKeyboardMarkup(keyboard_c)
            )
        except Exception as e:
            await ctx.bot.send_message(chat_id=chat_id, text=f"⚠️ Ошибка подтверждения: {e}")

    elif data.startswith("balance_confirm:"):
        try:
            amt = int(data.split(":")[1])
            rate = get_rate(chat_id)
            amt_str = f"{amt:,}".replace(",", " ") + " ₽"
            if rate:
                usdt_str = fmt_usd(amt / rate)
                rate_str = f" по курсу {rate} ₽"
            else:
                usdt_str = "— (курс не задан)"
                rate_str = ""
            keyboard_c = [
                [InlineKeyboardButton(f"✅ Подтвердить +{amt_str}", callback_data=f"balance_add:{amt}")],
                [InlineKeyboardButton("← Назад", callback_data="balance_topup")],
            ]
            await query.edit_message_text(
                f"💰 Пополнение баланса\n\nСумма: {amt_str}\nВ USDT: ≈ {usdt_str}{rate_str}\n\nПодтверди отправку:",
                reply_markup=InlineKeyboardMarkup(keyboard_c)
            )
        except Exception as e:
            await ctx.bot.send_message(chat_id=chat_id, text=f"⚠️ Ошибка подтверждения: {e}")

    elif data == "balance_topup_settle":
        cur_bal = balance.get(chat_id) or 0
        rate = get_rate(chat_id)
        if rate:
            bal_str = f"{round(cur_bal * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(cur_bal)
        else:
            bal_str = fmt_usd(cur_bal)
        hint = ""
        keyboard_b = [
            [InlineKeyboardButton("100к", callback_data="balance_confirm_settle:100000"),
             InlineKeyboardButton("150к", callback_data="balance_confirm_settle:150000")],
            [InlineKeyboardButton("200к", callback_data="balance_confirm_settle:200000"),
             InlineKeyboardButton("250к", callback_data="balance_confirm_settle:250000")],
            [InlineKeyboardButton("300к", callback_data="balance_confirm_settle:300000")],
            [InlineKeyboardButton("✏️ Своя сумма (₽)", callback_data="balance_custom_settle")],
            [InlineKeyboardButton("💵 Своя сумма (USDT)", callback_data="balance_custom_usdt_settle")],
            [InlineKeyboardButton("← Назад", callback_data="calculate")],
        ]
        await query.edit_message_text(
            f"💰 Текущий баланс: {bal_str}{hint}\n\nВыбери сумму пополнения:",
            reply_markup=InlineKeyboardMarkup(keyboard_b)
        )

    elif data.startswith("balance_add_settle:"):
        try:
            amt = int(data.split(":")[1])
            rate = get_rate(chat_id)
            if not rate:
                await query.edit_message_text("⚠️ Сначала установи курс через /setrate, потом пополняй.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В список", callback_data="go_list")]]))
                return
            amt_usd = amt / rate
            cur = balance.get(chat_id) or 0
            balance[chat_id] = cur + amt_usd
            topups.setdefault(chat_id, []).append({"date": datetime.now().strftime("%H:%M"), "amount": amt, "amount_usd": amt_usd, "rate": rate})
            new_bal = balance[chat_id]
            if rate:
                bal_str = f"{round(new_bal * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(new_bal)
            else:
                bal_str = fmt_usd(new_bal)
            amt_str = f"{amt:,}".replace(",", " ") + " ₽ (≈ " + fmt_usd(amt_usd) + ")"
            keyboard_b = [
                [InlineKeyboardButton("➕ Ещё пополнить", callback_data="balance_topup_settle")],
                [InlineKeyboardButton("← В список", callback_data="go_list")],
            ]
            await query.edit_message_text(
                f"✅ Баланс пополнен на {amt_str}\n\n💰 Новый баланс: {bal_str}",
                reply_markup=InlineKeyboardMarkup(keyboard_b)
            )
        except Exception as e:
            await ctx.bot.send_message(chat_id=chat_id, text=f"⚠️ Ошибка пополнения: {e}")

    elif data == "balance_topup":
        cur_bal = balance.get(chat_id) or 0
        rate = get_rate(chat_id)
        if rate:
            bal_str = f"{round(cur_bal * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(cur_bal)
        else:
            bal_str = fmt_usd(cur_bal)
        keyboard_b = [
            [InlineKeyboardButton("100к", callback_data="balance_confirm:100000"),
             InlineKeyboardButton("150к", callback_data="balance_confirm:150000")],
            [InlineKeyboardButton("200к", callback_data="balance_confirm:200000"),
             InlineKeyboardButton("250к", callback_data="balance_confirm:250000")],
            [InlineKeyboardButton("300к", callback_data="balance_confirm:300000")],
            [InlineKeyboardButton("✏️ Своя сумма (₽)", callback_data="balance_custom")],
            [InlineKeyboardButton("💵 Своя сумма (USDT)", callback_data="balance_custom_usdt")],
            [InlineKeyboardButton("← В список", callback_data="go_list")],
        ]
        await query.edit_message_text(
            f"💰 Текущий баланс: {bal_str}\n\nВыбери сумму пополнения:",
            reply_markup=InlineKeyboardMarkup(keyboard_b)
        )

    elif data.startswith("balance_add:"):
        try:
            amt = int(data.split(":")[1])
            rate = get_rate(chat_id)
            if not rate:
                await query.edit_message_text("⚠️ Сначала установи курс через /setrate, потом пополняй.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В список", callback_data="go_list")]]))
                return
            amt_usd = amt / rate
            cur = balance.get(chat_id) or 0
            balance[chat_id] = cur + amt_usd
            topups.setdefault(chat_id, []).append({"date": datetime.now().strftime("%H:%M"), "amount": amt, "amount_usd": amt_usd, "rate": rate})
            new_bal = balance[chat_id]
            if rate:
                bal_str = f"{round(new_bal * rate):,}".replace(",", " ") + " ₽ / " + fmt_usd(new_bal)
            else:
                bal_str = fmt_usd(new_bal)
            amt_str = f"{amt:,}".replace(",", " ") + " ₽ (≈ " + fmt_usd(amt_usd) + ")"
            keyboard_b = [
                [InlineKeyboardButton("➕ Ещё пополнить", callback_data="balance_topup")],
                [InlineKeyboardButton("← В список", callback_data="go_list")],
            ]
            await query.edit_message_text(
                f"✅ Баланс пополнен на {amt_str}\n\n💰 Новый баланс: {bal_str}",
                reply_markup=InlineKeyboardMarkup(keyboard_b)
            )
        except Exception as e:
            await ctx.bot.send_message(chat_id=chat_id, text=f"⚠️ Ошибка пополнения: {e}")

    elif data.startswith("edit_bank:"):
        msg_id = int(data.split(":")[1])
        ctx.user_data["editing_bank_msg_id"] = msg_id
        await query.edit_message_text(
            "Напиши название банка в чат:\n(Сбер, Тинькофф, ВТБ, Альфа...)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Игнорировать", callback_data="ignore_order")]])
        )

    elif data.startswith("force_add:"):
        msg_id = int(data.split(":")[1])
        parsed = ctx.user_data.pop(f"pending_{msg_id}", None)
        if not parsed:
            await query.edit_message_text("⚠️ Данные устарели, отправь заявку снова.")
            return
        order_id = next_id(chat_id)
        order = {
            "id": order_id, "label": "",
            "original_text": "",
            "bank": parsed.get("bank"), "amount": parsed.get("amount"),
            "currency": parsed.get("currency", "RUB"),
            "card": parsed.get("card"), "phone": parsed.get("phone"),
            "rate": get_rate(chat_id), "close_rate": None,
            "status": "open", "closed_by": None, "closed_at": None,
            "message_id": msg_id,
            "added_at": datetime.now().strftime("%H:%M"),
        }
        order["label"] = format_label(order)
        get_orders(chat_id).append(order)
        await query.edit_message_text(f"✅ Заявка #{order_id} добавлена без суммы:\n{order['label']}")

    elif data == "ignore_order":
        await query.edit_message_text("❌ Заявка проигнорирована.")
        # Удалить сообщение пользователя (исходное) и наш ответ через 10 сек
        bot_msg_id = query.message.message_id
        user_msg_id = query.message.reply_to_message.message_id if query.message.reply_to_message else None
        async def _cleanup():
            await asyncio.sleep(10)
            for mid in (bot_msg_id, user_msg_id):
                if mid:
                    try:
                        await ctx.bot.delete_message(chat_id=chat_id, message_id=mid)
                    except Exception:
                        pass
        asyncio.create_task(_cleanup())

    elif data.startswith("keep:"):
        order_id = int(data.split(":")[1])
        await query.edit_message_text(f"✅ Заявка #{order_id} оставлена как есть.")

    elif data.startswith("edit:"):
        order_id = int(data.split(":")[1])
        order = next((o for o in get_orders(chat_id) if o["id"] == order_id), None)
        if not order:
            await query.edit_message_text("⚠️ Заявка не найдена.")
            return
        missing = []
        if not order.get("bank"):
            missing.append("банк")
        if not order.get("amount"):
            missing.append("сумма")
        ctx.user_data["editing_order_id"] = order_id
        ctx.user_data["editing_chat_id"] = chat_id
        ctx.user_data["editing_missing"] = missing
        fields = " и ".join(missing)
        await query.edit_message_text(
            f"✏️ Заявка #{order_id}\nНапиши {fields} в чат\n\nПример: Сбер 15000",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← Меню", callback_data=f"keep:{order_id}")]])
        )

    elif data == "show_logs":
        all_sessions = sessions.get(chat_id, [])
        if not all_sessions:
            await query.answer("Нет завершённых сессий", show_alert=True)
            return
        lines = []
        for i, s in enumerate(all_sessions, 1):
            lines.append(f"━━━ Сессия #{i} ({s['date']}) ━━━")
            usd_total = 0
            if s.get("orders"):
                for o in s["orders"]:
                    amt = o.get("amount")
                    if not amt:
                        lines.append("  ?")
                        continue
                    cur = o.get("currency", "RUB")
                    r = o.get("close_rate") or o.get("rate")
                    amt_str = f"{int(amt):,}".replace(",", " ") if float(amt).is_integer() else str(amt)
                    rate_str = f" × {r}" if r and cur == "RUB" else ""
                    req = ""
                    if o.get("card"):
                        req = f" · ···{o['card'][-4:]}"
                    elif o.get("phone"):
                        req = f" · ···{o['phone'][-4:]}"
                    bank = o.get("bank") or "?"
                    line_usd = ""
                    if cur in ("USDT", "USD", "USDC"):
                        usd_total += amt
                    elif cur == "RUB" and r:
                        order_usd = amt / r
                        usd_total += order_usd
                        line_usd = f" → {fmt_usd(order_usd)}"
                    lines.append(f"  {bank} · {amt_str} {cur}{rate_str}{req}{line_usd}")
            lines.append(f"  Заявок: {s.get('count', '?')}")
            if usd_total > 0:
                lines.append(f"  Итого: ≈ {fmt_usd(usd_total)}")
            if s.get("paid"):
                lines.append(f"  Оплачено: {fmt_usd(s['paid'])}")
            if s.get("debt"):
                lines.append(f"  Долг: {fmt_usd(s['debt'])}")
            lines.extend(_format_topups_block(s))
            lines.append("")
        cur_debt = debt.get(chat_id, 0)
        if cur_debt > 0:
            lines.append(f"⚠️ Текущий долг: {fmt_usd(cur_debt)}")
        text = "📒 История сессий\n\n" + "\n".join(lines)
        if len(text) > 4000:
            text = text[:3990] + "\n..."
        await query.edit_message_text(text,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))

    elif data == "del_menu":
        # Меню выбора заявки для удаления
        open_orders = [o for o in get_orders(chat_id) if o["status"] == "open"]
        if not open_orders:
            await query.answer("Нет открытых заявок", show_alert=True)
            return
        kb = [[InlineKeyboardButton(f"#{o['id']}  {o['label']}", callback_data=f"del_ask:{o['id']}")] for o in open_orders]
        kb.append([InlineKeyboardButton("← Назад", callback_data="go_list")])
        await query.edit_message_text(
            "🗑 *Какую заявку удалить?*\n━━━━━━━━━━━━━━━━━━━━\nНажми на заявку для удаления:",
            reply_markup=InlineKeyboardMarkup(kb),
            parse_mode="Markdown"
        )

    elif data.startswith("del_ask:"):
        # Подтверждение удаления
        order_id = int(data.split(":")[1])
        order = next((o for o in get_orders(chat_id) if o["id"] == order_id), None)
        if not order:
            await query.answer("Заявка не найдена", show_alert=True)
            return
        if order["status"] != "open":
            await query.answer("Заявку нельзя удалить (уже закрыта или удалена)", show_alert=True)
            return
        text = (
            f"⚠️ *Удалить заявку #{order_id}?*\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"{order['label']}\n"
            f"Добавлена: {order['added_at']}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"Это действие нельзя будет отменить."
        )
        kb = [[
            InlineKeyboardButton("🗑 Да, удалить", callback_data=f"del_confirm:{order_id}"),
            InlineKeyboardButton("← Отмена", callback_data="del_menu"),
        ]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")

    elif data.startswith("del_confirm:"):
        # Собственно удаление
        order_id = int(data.split(":")[1])
        order = next((o for o in get_orders(chat_id) if o["id"] == order_id), None)
        if not order:
            await query.answer("Заявка не найдена", show_alert=True)
            return
        if order["status"] != "open":
            await query.answer("Заявку нельзя удалить", show_alert=True)
            return
        order["status"] = "deleted"
        # Если в сообщении была только эта одна заявка — удаляем сообщение
        msg_deleted = await _try_delete_source_message(
            ctx, chat_id, order["message_id"], get_orders(chat_id)
        )
        if not msg_deleted:
            try:
                await ctx.bot.set_message_reaction(
                    chat_id=chat_id, message_id=order["message_id"],
                    reaction=[ReactionTypeEmoji(emoji="👎")]
                )
            except Exception:
                pass
        suffix = "\n_Исходное сообщение удалено_" if msg_deleted else ""
        await query.edit_message_text(
            f"🗑 *Заявка #{order_id} удалена*\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"{order['label']}{suffix}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]),
            parse_mode="Markdown"
        )

    elif data in ("back_list", "cancel_conv", "go_list"):
        ctx.user_data.pop("waiting_close_rate", None)
        await _show_list_inline(query, ctx, chat_id)

async def report(update, ctx):
    chat_id = update.effective_chat.id
    all_orders = [o for o in get_orders(chat_id) if o["status"] in ("open", "closed")]
    if not all_orders:
        await update.message.reply_text("📭 Заявок нет.")
        return

    open_o   = [o for o in all_orders if o["status"] == "open"]
    closed_o = [o for o in all_orders if o["status"] == "closed"]
    rate     = get_rate(chat_id)

    # Список заявок
    lines = []
    for o in all_orders:
        if o["status"] == "open":
            lines.append(f"⏳ #{o['id']} {o['label']}")
        elif o["status"] == "closed":
            rate_str = f" (курс {o['close_rate']})" if o.get("close_rate") else ""
            lines.append(f"✅ #{o['id']} {o['label']}{rate_str}")

    # Суммы закрытых
    totals_rub = 0
    totals_usdt = 0
    usd_total = 0
    for o in closed_o:
        if not o.get("amount"):
            continue
        cur = o.get("currency", "RUB")
        r = o.get("close_rate") or rate
        if cur in ("USDT", "USD", "USDC"):
            totals_usdt += o["amount"]
            usd_total += o["amount"]
        elif cur == "RUB":
            totals_rub += o["amount"]
            if r:
                usd_total += o["amount"] / r

    summary = []
    if totals_rub:
        summary.append(f"  RUB: {int(totals_rub):,}".replace(",", " "))
    if totals_usdt:
        summary.append(f"  USDT: {fmt_usd(totals_usdt).replace(' USD', '')}")
    if usd_total and rate:
        summary.append(f"  Итого ≈ {fmt_usd(usd_total)}")

    rate_text = f"💱 Курс: {rate} ₽\n" if rate else ""
    text = (
        f"📊 Отчёт за {datetime.now().strftime('%d.%m.%Y')}\n"
        f"{rate_text}\n"
        + "\n".join(lines)
        + f"\n\n📦 Всего: {len(all_orders)} | ⏳ Открыто: {len(open_o)} | ✅ Закрыто: {len(closed_o)}"
    )
    if summary:
        text += "\n\n💰 Сумма закрытых:\n" + "\n".join(summary)

    await update.message.reply_text(text)



async def cmd_calculate(update, ctx):
    chat_id = update.effective_chat.id
    closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]

    if not closed_orders:
        await update.message.reply_text("📭 Нет выполненных заявок для подсчёта.")
        return

    rate = get_rate(chat_id)
    totals = {}
    usd_total = 0

    for o in closed_orders:
        if not o.get("amount"):
            continue
        cur = o.get("currency", "RUB")
        totals[cur] = totals.get(cur, 0) + o["amount"]
        r = o.get("close_rate") or rate
        if cur in ("USDT", "USD", "USDC"):
            usd_total += o["amount"]
        elif cur == "RUB" and r:
            usd_total += o["amount"] / r

    sum_lines = ["\n".join([f"  {int(v):,} {k}".replace(",", " ") for k, v in totals.items()])]
    usd_line = f"\n💵 Итого ≈ {fmt_usd(usd_total)}" if usd_total > 0 else ""
    if not usd_total and not rate:
        usd_line = "\n💡 Установи курс (/setrate) для подсчёта в USD"

    text = (
        f"🧾 Итог сессии ({len(closed_orders)} заявок):\n\n"
        + "\n".join(sum_lines)
        + usd_line
        + "\n\nНажми «Посчитались» чтобы обнулить и начать новую сессию."
    )
    keyboard = [[
        InlineKeyboardButton("✅ Посчитались", callback_data="settle"),
        InlineKeyboardButton("← Меню", callback_data="back_list"),
    ]]
    await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

async def cmd_del(update, ctx):
    """Удалить заявку — открывает меню выбора, либо удаляет по reply"""
    msg = update.message
    chat_id = msg.chat_id
    if not msg.reply_to_message:
        open_orders = [o for o in get_orders(chat_id) if o["status"] == "open"]
        if not open_orders:
            await msg.reply_text(
                "📭 *Нет открытых заявок для удаления*",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]),
                parse_mode="Markdown"
            )
            return
        kb = [[InlineKeyboardButton(f"#{o['id']}  {o['label']}", callback_data=f"del_ask:{o['id']}")] for o in open_orders]
        kb.append([InlineKeyboardButton("← Назад", callback_data="go_list")])
        await msg.reply_text(
            "🗑 *Какую заявку удалить?*\n━━━━━━━━━━━━━━━━━━━━\nНажми на заявку для удаления:",
            reply_markup=InlineKeyboardMarkup(kb),
            parse_mode="Markdown"
        )
        return
    reply_msg_id = msg.reply_to_message.message_id
    order = next((o for o in get_orders(chat_id) if o["message_id"] == reply_msg_id), None)
    if not order:
        await msg.reply_text("⚠️ Заявка не найдена.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    if order["status"] == "closed":
        await msg.reply_text("⚠️ Заявка уже закрыта. Сначала отмени закрытие через /list.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    if order["status"] == "settled":
        await msg.reply_text("⚠️ Заявка уже посчитана и не может быть удалена.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    order["status"] = "deleted"
    msg_deleted = await _try_delete_source_message(ctx, chat_id, reply_msg_id, get_orders(chat_id))
    if not msg_deleted:
        try:
            await ctx.bot.set_message_reaction(chat_id=chat_id, message_id=reply_msg_id, reaction=[ReactionTypeEmoji(emoji="👎")])
        except Exception:
            pass
    suffix = "\n_Исходное сообщение удалено_" if msg_deleted else ""
    await msg.reply_text(
        f"🗑 Заявка #{order['id']} удалена.\n{order['label']}{suffix}",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]),
        parse_mode="Markdown"
    )

async def cmd_refresh(update, ctx):
    """Обновить заявку — перечитать изменённое сообщение"""
    msg = update.message
    if not msg.reply_to_message:
        await msg.reply_text("⚠️ Ответь на сообщение с заявкой чтобы обновить её.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    chat_id = msg.chat_id
    reply_msg = msg.reply_to_message
    reply_msg_id = reply_msg.message_id
    order = next((o for o in get_orders(chat_id) if o["message_id"] == reply_msg_id), None)
    if not order:
        await msg.reply_text("⚠️ Заявка не найдена.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    if order["status"] != "open":
        await msg.reply_text("⚠️ Можно обновить только открытую заявку.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    new_text = reply_msg.text
    if not new_text:
        await msg.reply_text("⚠️ Не могу прочитать текст сообщения.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    parsed = await parse_with_groq(new_text)
    if not parsed:
        await msg.reply_text("⚠️ Не смог распознать заявку в сообщении.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    # Постобработка
    if parsed.get("card") and not parsed.get("phone"):
        digits = re.sub(r"\D", "", str(parsed["card"]))
        if 10 <= len(digits) <= 11:
            parsed["phone"] = "+7" + digits[-10:]
            parsed["card"] = None
    if not parsed.get("bank"):
        bank_map = {
            "сбер": "Сбер", "тинь": "Тинькофф", "тинк": "Тинькофф", "тинькофф": "Тинькофф",
            "втб": "ВТБ", "альфа": "Альфа", "озон": "Озон", "райф": "Райффайзен",
            "газпром": "Газпромбанк", "совком": "Совкомбанк", "отп": "ОТП",
            "мкб": "МКБ", "псб": "ПСБ", "юмани": "ЮMoney", "qiwi": "QIWI",
            "почта": "Почта Банк", "уралсиб": "Уралсиб", "хоум": "Хоум Кредит",
            "росбанк": "Росбанк", "открытие": "Открытие", "яндекс": "Яндекс Банк",
        }
        for key, name in bank_map.items():
            if key in new_text.lower():
                parsed["bank"] = name
                break
    # Обновляем заявку
    order["bank"] = parsed.get("bank", order["bank"])
    order["amount"] = parse_amount(parsed.get("amount")) or order["amount"]
    order["currency"] = parsed.get("currency", order["currency"])
    order["card"] = parsed.get("card", order["card"])
    order["phone"] = parsed.get("phone", order["phone"])
    order["label"] = format_label(order)
    await msg.reply_text(f"✅ Заявка #{order['id']} обновлена:\n{order['label']}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))

async def _send_list_to_chat(ctx, chat_id):
    """Отправляет в чат свежее меню /list, удаляя предыдущее. Используется автообновлением."""
    try:
        open_orders = [o for o in get_orders(chat_id) if o["status"] == "open"]
        closed_orders = [o for o in get_orders(chat_id) if o["status"] == "closed"]
        settled_orders = [o for o in get_orders(chat_id) if o["status"] == "settled"]
        rate_text = f"💱 Курс: {current_rate[chat_id]} ₽\n" if chat_id in current_rate else ""
        has_debt = debt.get(chat_id, 0) > 0
        has_sessions = bool(sessions.get(chat_id))
        if not open_orders and not closed_orders and not settled_orders and not has_debt and not has_sessions:
            text = f"{rate_text}{_balance_line(chat_id)}📭 Активных заявок нет. Отправляй заявки в чат."
            keyboard = [[InlineKeyboardButton("💰 Пополнить баланс", callback_data="balance_topup")]]
        else:
            keyboard = [[InlineKeyboardButton(f"#{o['id']}  {o['label']}  ({o['added_at']})", callback_data=f"ask:{o['id']}")] for o in open_orders]
            btn_row = []
            if closed_orders or settled_orders:
                btn_row.append(InlineKeyboardButton("💰 Результат дня", callback_data="show_total"))
            if closed_orders or (balance.get(chat_id) is not None and balance[chat_id] < 0):
                btn_row.append(InlineKeyboardButton("🧾 Посчитаться", callback_data="calculate"))
            if btn_row:
                keyboard.append(btn_row)
            if sessions.get(chat_id):
                keyboard.append([InlineKeyboardButton("📒 Логи сессий", callback_data="show_logs")])
            last_closed = next((o for o in reversed(get_orders(chat_id)) if o["status"] == "closed"), None)
            if last_closed:
                keyboard.append([InlineKeyboardButton(f"↩️ Отменить закрытие #{last_closed['id']}", callback_data=f"undo:{last_closed['id']}")])
            _bulk = last_bulk_close.get(chat_id) or []
            _still = [oid for oid in _bulk if any(o["id"]==oid and o["status"]=="closed" for o in get_orders(chat_id))]
            if len(_still) >= 2:
                keyboard.append([InlineKeyboardButton(f"↩️ Отменить закрытие всех ({len(_still)})", callback_data="undo_all")])
            keyboard.append([InlineKeyboardButton("💰 Пополнить баланс", callback_data="balance_topup")])
            if open_orders:
                keyboard.append([InlineKeyboardButton(f"✅ Закрыть все ({len(open_orders)})", callback_data="close_all")])
                keyboard.append([InlineKeyboardButton("🗑 Удалить заявку", callback_data="del_menu")])
            text = rate_text + _balance_line(chat_id) + (f"📋 Активные заявки — {len(open_orders)} шт.\nНажми чтобы закрыть:" if open_orders else "📭 Активных заявок нет.")
        markup = InlineKeyboardMarkup(keyboard) if keyboard else None
        prev_id = last_list_msg.get(chat_id)
        if prev_id:
            try:
                await ctx.bot.delete_message(chat_id=chat_id, message_id=prev_id)
            except Exception:
                pass
            try:
                if chat_id in bot_msgs and prev_id in bot_msgs[chat_id]:
                    bot_msgs[chat_id].remove(prev_id)
            except Exception:
                pass
        sent = await ctx.bot.send_message(chat_id=chat_id, text=text, reply_markup=markup)
        last_list_msg[chat_id] = sent.message_id
    except Exception as e:
        logger.warning(f"_send_list_to_chat: {e}")

async def handle_reaction(update, ctx):
    """Если пользователь поставил 👍 на любое сообщение — открыть /list."""
    react = update.message_reaction
    if not react:
        return
    user = getattr(react, "user", None)
    if not user or getattr(user, "is_bot", False):
        return
    def _emojis(rs):
        out = set()
        for r in (rs or []):
            e = getattr(r, "emoji", None)
            if e:
                out.add(e)
        return out
    new_e = _emojis(react.new_reaction)
    old_e = _emojis(react.old_reaction)
    if "👍" in new_e and "👍" not in old_e:
        await _send_list_to_chat(ctx, react.chat.id)

async def cmd_logs(update, ctx):
    """Показать историю всех сессий"""
    chat_id = update.effective_chat.id
    all_sessions = sessions.get(chat_id, [])
    if not all_sessions:
        await update.message.reply_text("📭 Нет завершённых сессий.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))
        return
    lines = []
    for i, s in enumerate(all_sessions, 1):
        lines.append(f"━━━ Сессия #{i} ({s['date']}) ━━━")
        usd_total = 0
        if s.get("orders"):
            for o in s["orders"]:
                amt = o.get("amount")
                if not amt:
                    lines.append("  ?")
                    continue
                cur = o.get("currency", "RUB")
                r = o.get("close_rate") or o.get("rate")
                amt_str = f"{int(amt):,}".replace(",", " ") if float(amt).is_integer() else str(amt)
                rate_str = f" × {r}" if r and cur == "RUB" else ""
                req = ""
                if o.get("card"):
                    req = f" · ···{o['card'][-4:]}"
                elif o.get("phone"):
                    req = f" · ···{o['phone'][-4:]}"
                bank = o.get("bank") or "?"
                line_usd = ""
                if cur in ("USDT", "USD", "USDC"):
                    usd_total += amt
                elif cur == "RUB" and r:
                    order_usd = amt / r
                    usd_total += order_usd
                    line_usd = f" → {fmt_usd(order_usd)}"
                lines.append(f"  {bank} · {amt_str} {cur}{rate_str}{req}{line_usd}")
        lines.append(f"  Заявок: {s.get('count', '?')}")
        if usd_total > 0:
            lines.append(f"  Итого: ≈ {fmt_usd(usd_total)}")
        if s.get("paid"):
            lines.append(f"  Оплачено: {fmt_usd(s['paid'])}")
        if s.get("debt"):
            lines.append(f"  Долг: {fmt_usd(s['debt'])}")
        lines.extend(_format_topups_block(s))
        lines.append("")
    cur_debt = debt.get(chat_id, 0)
    if cur_debt > 0:
        lines.append(f"⚠️ Текущий долг: {fmt_usd(cur_debt)}")
    text = "📒 История сессий\n\n" + "\n".join(lines)
    if len(text) > 4000:
        text = text[:3990] + "\n..."
    await update.message.reply_text(text,
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("← В меню", callback_data="go_list")]]))

async def cancel_conv(update, ctx):
    await update.message.reply_text("↩️ Отменено.")
    return ConversationHandler.END

async def cmd_stop(update, ctx):
    chat_id = update.effective_chat.id
    bot_active[chat_id] = False
    await update.message.reply_text(
        "⏸ *БОТ НА ПАУЗЕ* ⏸\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "🛑 Приём новых заявок остановлен\n"
        "💾 Текущие данные сохранены\n\n"
        "▶️ Чтобы возобновить — /start",
        parse_mode="Markdown"
    )

async def cmd_start_active(update, ctx):
    chat_id = update.effective_chat.id
    was_stopped = not bot_active.get(chat_id, True)
    bot_active[chat_id] = True

    rate = current_rate.get(chat_id)
    open_count = sum(1 for o in get_orders(chat_id) if o["status"] == "open")
    closed_count = sum(1 for o in get_orders(chat_id) if o["status"] in ("closed", "settled"))

    if was_stopped:
        text = (
            "🔥 *БОТ СНОВА В ИГРЕ* 🔥\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            "✅ Приём заявок возобновлён\n"
            "🚀 Готов парсить реквизиты\n\n"
        )
        if rate:
            text += f"💱 Текущий курс: *{rate} ₽*\n"
        if open_count or closed_count:
            text += f"📋 Заявок открыто: *{open_count}*  ·  закрыто: *{closed_count}*\n"
        text += (
            "\n━━━━━━━━━━━━━━━━━━━━\n"
            "📌 /list — активные заявки\n"
            "📊 /report — отчёт за день"
        )
        await update.message.reply_text(text, parse_mode="Markdown")
    else:
        text = (
            "✨ *EXCHANGE TRACKER BOT* ✨\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "🤖 Умный парсер платёжных заявок\n"
            "⚡️ На базе AI · мгновенное распознавание\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "💼 *ЧТО Я УМЕЮ*\n\n"
            "🎯 Распознаю заявки в любом формате:\n"
            "    `Сбер 25к +79161234567`\n"
            "    `2202 2017 4018 6792 — 10000 ВТБ`\n"
            "    `2 платежа по 45000`\n\n"
            "📦 Парсю списки заявок в одном сообщении\n"
            "💱 Считаю итоги в RUB и USD по курсу\n"
            "📒 Веду историю сессий\n"
            "👍 Ставлю реакцию на распознанные заявки\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            "🎮 *КОМАНДЫ*\n\n"
            "💱 /setrate 77   — установить курс\n"
            "📋 /list         — активные заявки\n"
            "📊 /report       — отчёт за день\n"
            "📒 /logs         — история сессий\n"
            "✏️ /refresh      — перечитать заявку (ответом)\n"
            "🗑 /del          — удалить заявку (ответом)\n"
            "⏸ /stop          — приостановить приём\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
        )
        if rate:
            text += f"💱 *Курс установлен:*  {rate} ₽\n"
        else:
            text += "⚠️ *Курс не установлен* — задай через /setrate\n"
        if open_count or closed_count:
            text += f"📊 *Сейчас:*  открыто {open_count}  ·  закрыто {closed_count}\n"
        text += (
            "\n💡 *Просто пиши заявки в чат* — я их распознаю автоматически!\n"
            "🚀 Поехали!"
        )
        await update.message.reply_text(text, parse_mode="Markdown")

def main():
    app = (
        Application.builder()
        .token(TELEGRAM_TOKEN)
        .read_timeout(30)
        .write_timeout(30)
        .connect_timeout(30)
        .pool_timeout(30)
        .build()
    )
    rate_conv = ConversationHandler(
        entry_points=[CommandHandler("setrate", setrate_start)],
        states={WAITING_RATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, setrate_save)]},
        fallbacks=[CommandHandler("cancel", cancel_conv)]
    )
    app.add_handler(CommandHandler("start", cmd_start_active))
    app.add_handler(CommandHandler("stop", cmd_stop))
    app.add_handler(CommandHandler("del", cmd_del))
    app.add_handler(CommandHandler("logs", cmd_logs))
    app.add_handler(CommandHandler("refresh", cmd_refresh))
    app.add_handler(rate_conv)
    app.add_handler(CommandHandler("list", show_list))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageReactionHandler(handle_reaction))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # === Авто-очистка: храним только последние 3 сообщения бота, остальные удаляем ===
    from telegram import Bot as _TGBot
    if not getattr(_TGBot, "_send_patched", False):
        _orig_send = _TGBot.send_message
        async def _tracked_send_message(self, chat_id, *args, **kwargs):
            msg = await _orig_send(self, chat_id, *args, **kwargs)
            try:
                ml = bot_msgs.setdefault(chat_id, [])
                ml.append(msg.message_id)
                while len(ml) > 3:
                    old_id = ml.pop(0)
                    try:
                        await self.delete_message(chat_id=chat_id, message_id=old_id)
                    except Exception:
                        pass
            except Exception:
                pass
            return msg
        _TGBot.send_message = _tracked_send_message
        _TGBot._send_patched = True
    print("✅ Бот запущен...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
